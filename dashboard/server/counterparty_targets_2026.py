import logging
import threading
from pathlib import Path
from typing import Any, Dict, Tuple

import openpyxl

RESOURCE_PATH = Path(__file__).parent / "resources" / "counterparty_targets_2026.xlsx"

_CACHE_LOCK = threading.Lock()
_CACHE: Dict[str, Any] = {"mtime": None, "offline": {}, "online": {}, "meta": {}}


def _norm_min(val: Any) -> str:
    if val is None:
        return ""
    return str(val).replace("\u00A0", " ").strip()


def _normalize_name(val: Any) -> str:
    return _norm_min(val)


def _normalize_upper(val: Any) -> str:
    text = _norm_min(val)
    if not text or text in {"-", "–", "—"}:
        return "미입력"
    return text


def _parse_value(val: Any) -> float | None:
    try:
        num = float(val)
        return num * 1e8
    except Exception:
        return None


def _load_sheet(ws, value_col: str) -> Tuple[Dict[Tuple[str, str], float], Dict[Tuple[str, str], Dict[str, Any]]]:
    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    headers = next(rows, None)
    if not headers:
        logging.warning("[counterparty_targets_2026] sheet=%s empty header", ws.title)
        return {}, {}
    header_idx = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None and str(h).strip()}
    required = ["기업명", "카운터파티", value_col]
    if not all(col in header_idx for col in required):
        logging.warning("[counterparty_targets_2026] sheet=%s missing columns %s", ws.title, required)
        return {}, {}

    org_id_col: str | None = None
    for name in header_idx.keys():
        lowered = name.lower()
        if lowered in {"orgid", "org_id"} or name in {"기업ID", "기업 Id", "기업 id"}:
            org_id_col = name
            break

    entries: Dict[Tuple[str, str], list] = {}
    meta: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        org_raw = row[header_idx["기업명"]] if header_idx["기업명"] < len(row) else None
        upper_raw = row[header_idx["카운터파티"]] if header_idx["카운터파티"] < len(row) else None
        val_raw = row[header_idx[value_col]] if header_idx[value_col] < len(row) else None
        org_id_raw = row[header_idx[org_id_col]] if org_id_col and header_idx[org_id_col] < len(row) else None

        org = _normalize_name(org_raw)
        upper = _normalize_upper(upper_raw)
        key = (org, upper)
        org_id = _normalize_name(org_id_raw)

        val = _parse_value(val_raw)
        if val is None:
            logging.warning(
                "[counterparty_targets_2026] invalid value sheet=%s row=%d key=(%s|%s) raw=%s",
                ws.title,
                excel_row_idx,
                org,
                upper,
                val_raw,
            )
            continue
        entries.setdefault(key, []).append((excel_row_idx, val_raw, val))

    result: Dict[Tuple[str, str], float] = {}
    for key, items in entries.items():
        if len(items) > 1:
            rows = [it[0] for it in items]
            values = [it[1] for it in items]
            logging.warning(
                "[counterparty_targets_2026] DUPLICATE_KEY sheet=%s key=(%s|%s) rows=%s values=%s",
                ws.title,
                key[0],
                key[1],
                rows,
                values,
            )
            continue
        result[key] = float(items[0][2])
        if key not in meta:
            meta[key] = {
                "sheet": ws.title,
                "row": items[0][0],
                "orgId": org_id or None,
                "orgName": key[0],
                "upperOrg": key[1],
                "orgRaw": org_raw,
                "upperRaw": upper_raw,
            }
    return result, meta


def load_counterparty_targets_2026() -> Tuple[Dict[Tuple[str, str], float], Dict[Tuple[str, str], float], Dict[Tuple[str, str], Dict[str, Any]], str]:
    """
    Returns offline_targets, online_targets (won unit) and version string based on xlsx mtime.
    """
    with _CACHE_LOCK:
        try:
            stat = RESOURCE_PATH.stat()
        except FileNotFoundError:
            logging.warning("[counterparty_targets_2026] resource not found: %s", RESOURCE_PATH)
            return {}, {}, {}, "xlsx_mtime:none"
        mtime = int(stat.st_mtime)
        if _CACHE.get("mtime") == mtime:
            return _CACHE["offline"], _CACHE["online"], _CACHE["meta"], f"xlsx_mtime:{mtime}"

        wb = openpyxl.load_workbook(RESOURCE_PATH, data_only=True, read_only=True)
        offline_ws = wb["26 출강 타겟"] if "26 출강 타겟" in wb.sheetnames else None
        online_ws = wb["26 온라인 타겟"] if "26 온라인 타겟" in wb.sheetnames else None

        offline, offline_meta = _load_sheet(offline_ws, "26 출강 타겟") if offline_ws else ({}, {})
        online, online_meta = _load_sheet(online_ws, "26 온라인 타겟") if online_ws else ({}, {})

        merged_meta: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for key, meta in offline_meta.items():
            merged_meta.setdefault(key, {"orgId": meta.get("orgId"), "orgName": meta.get("orgName"), "upperOrg": meta.get("upperOrg"), "hasOffline": False, "hasOnline": False, "sheet": meta.get("sheet"), "row": meta.get("row")})
            merged_meta[key]["hasOffline"] = True
        for key, meta in online_meta.items():
            existing = merged_meta.setdefault(key, {"orgId": meta.get("orgId"), "orgName": meta.get("orgName"), "upperOrg": meta.get("upperOrg"), "hasOffline": False, "hasOnline": False, "sheet": meta.get("sheet"), "row": meta.get("row")})
            if not existing.get("orgId") and meta.get("orgId"):
                existing["orgId"] = meta.get("orgId")
            if not existing.get("sheet"):
                existing["sheet"] = meta.get("sheet")
            if not existing.get("row"):
                existing["row"] = meta.get("row")
            existing["hasOnline"] = True

        _CACHE.update({"mtime": mtime, "offline": offline, "online": online, "meta": merged_meta})
        return offline, online, merged_meta, f"xlsx_mtime:{mtime}"
