import os
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import PlainTextResponse, Response
from io import BytesIO
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import json
from typing import Any

from . import database as db
from .json_compact import compact_won_groups_json
from .markdown_compact import won_groups_compact_to_markdown
from .statepath_engine import build_statepath
from .report_scheduler import run_daily_counterparty_risk_job, get_cached_report, _load_status
from .llm_target_attainment import (
    TargetAttainmentRequest,
    run_target_attainment,
    validate_payload_limits,
    MAX_TARGET_ATTAINMENT_REQUEST_BYTES,
)
from .agents.daily_report_v2.orchestrator import run_pipeline as run_daily_report_v2_pipeline

router = APIRouter(prefix="/api")

# TODO: remove _debug/won-json-runtime after won-json memo.get investigation is resolved.
if os.getenv("DEBUG_WON_JSON") == "1":

    @router.get("/_debug/won-json-runtime")
    def debug_won_json_runtime() -> dict:
        import dashboard.server.database as db_mod
        from pathlib import Path

        db_file = Path(db_mod.__file__).resolve()
        text = db_file.read_text(encoding="utf-8", errors="ignore")
        occurrences = [idx + 1 for idx, line in enumerate(text.splitlines()) if "created_at_ts" in line][:20]
        return {
            "db_file": str(db_file),
            "db_mtime": db_file.stat().st_mtime,
            "has_memo_get_createdAt": 'memo.get("createdAt")' in text,
            "created_at_ts_occurrences": occurrences,
        }


@router.get("/sizes")
def get_sizes() -> dict:
    """
    Distinct organization sizes.
    """
    try:
        return {"sizes": db.list_sizes()}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs")
def get_organizations(
    size: str = Query("전체", description='조직 규모 필터 (예: "대기업", "전체")'),
    search: str | None = Query(None, description="조직명 검색어"),
    limit: int = Query(200, ge=1, le=500, description="최대 반환 수"),
    offset: int = Query(0, ge=0, description="시작 offset"),
) -> dict:
    try:
        items = db.list_organizations(size=size, search=search, limit=limit, offset=offset)
        return {"items": items}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}/memos")
def get_org_memos(org_id: str, limit: int = Query(100, ge=1, le=500)) -> dict:
    try:
        return {"items": db.get_org_memos(org_id=org_id, limit=limit)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}/people")
def get_org_people(
    org_id: str,
    has_deal: bool | None = Query(None, alias="hasDeal", description="딜 여부 필터"),
) -> dict:
    try:
        people = db.get_people_for_org(org_id=org_id, has_deal=has_deal)
        return {"items": people}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/people/{person_id}/deals")
def get_person_deals(person_id: str) -> dict:
    try:
        return {"items": db.get_deals_for_person(person_id=person_id)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/people/{person_id}/memos")
def get_person_memos(person_id: str, limit: int = Query(200, ge=1, le=500)) -> dict:
    try:
        return {"items": db.get_memos_for_person(person_id=person_id, limit=limit)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/deals/{deal_id}/memos")
def get_deal_memos(deal_id: str, limit: int = Query(200, ge=1, le=500)) -> dict:
    try:
        return {"items": db.get_memos_for_deal(deal_id=deal_id, limit=limit)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/deal-check")
def get_deal_check(team: str = Query(..., description="팀 키 (edu1|edu2)")) -> dict:
    try:
        return {"items": db.get_deal_check(team)}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/deal-check/edu1")
def get_edu1_deal_check() -> dict:
    try:
        return {"items": db.get_deal_check("edu1")}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/deal-check/edu2")
def get_edu2_deal_check() -> dict:
    try:
        return {"items": db.get_deal_check("edu2")}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/ops/2026-online-retention")
def get_ops_2026_online_retention() -> dict:
    try:
        return db.get_ops_2026_online_retention()
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/qc/deal-errors/summary")
def get_qc_deal_errors_summary(team: str = Query("all", description="all|edu1|edu2|public")) -> dict:
    try:
        return db.get_qc_deal_errors_summary(team=team)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/qc/deal-errors/person")
def get_qc_deal_errors_for_owner(
    owner: str = Query(..., description="담당자 이름"), team: str = Query("all", description="all|edu1|edu2|public")
) -> dict:
    try:
        return db.get_qc_deal_errors_for_owner(team=team, owner=owner)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/qc/monthly-revenue-report")
def get_qc_monthly_revenue_report(
    team: str = Query(..., description="edu1|edu2|public"),
    year: int = Query(..., ge=2000, le=2100, description="연도 (YYYY)"),
    month: int = Query(..., ge=1, le=12, description="월 (1-12)"),
    history_from: str | None = Query(None, description="선택 월까지 포함할 과거 시작 월(YYYY-MM)"),
) -> dict:
    try:
        return db.get_qc_monthly_revenue_report(team=team, year=year, month=month, history_from=history_from)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/qc/monthly-revenue-report/xlsx")
def download_qc_monthly_revenue_report_xlsx(
    team: str = Query(..., description="edu1|edu2|public"),
    year: int = Query(..., ge=2000, le=2100, description="연도 (YYYY)"),
    month: int = Query(..., ge=1, le=12, description="월 (1-12)"),
) -> Response:
    try:
        data = db.get_qc_monthly_revenue_report(team=team, year=year, month=month)
        items = data.get("reportDeals", []) or []

        wb = Workbook()
        ws = wb.active
        ws.title = "매출신고"

        headers = ["코스 ID", "이름", "담당자", "상태", "계약 체결일", "금액(원)", "수강시작일", "수강종료일"]
        ws.append(headers)

        try:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
        except Exception:
            pass

        for row in items:
            owners = row.get("owners") or ""
            if isinstance(owners, list):
                owners = ", ".join(owners)
            ws.append(
                [
                    row.get("courseId") or "",
                    row.get("dealName") or "",
                    owners or "",
                    row.get("status") or "",
                    row.get("contractDate") or "",
                    row.get("amount") if row.get("amount") is not None else "",
                    row.get("startDate") or "",
                    row.get("endDate") or "",
                ]
            )

        try:
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=6)
                if isinstance(c.value, (int, float)):
                    c.number_format = "#,##0"
        except Exception:
            pass

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        team_label = getattr(db, "QC_TEAM_LABELS", {}).get(team, team)
        mm = f"{month:02d}"
        filename = f"{team_label}_{year}년_{mm}월_매출신고.xlsx"
        quoted = quote(filename)
        ascii_fallback = f"{team}_{year}_{mm}_revenue.xlsx"
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f'attachment; filename="{ascii_fallback}"; filename*=UTF-8\'\'{quoted}',
            "Cache-Control": "no-store",
        }
        return Response(content=bio.getvalue(), media_type=headers["Content-Type"], headers=headers)

    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/rank/2025-deals")
def get_rank_2025_deals(
    size: str = Query("전체", description='조직 규모 필터 (예: "대기업", "전체")')
) -> dict:
    try:
        return {"items": db.get_rank_2025_deals(size=size)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/rank/mismatched-deals")
def get_rank_mismatched_deals(
    size: str = Query("대기업", description='조직 규모 필터 (예: "대기업", "전체")')
) -> dict:
    try:
        return {"items": db.get_mismatched_deals(size=size)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/rank/won-yearly-totals")
def get_rank_won_yearly_totals() -> dict:
    try:
        return {"items": db.get_won_totals_by_size()}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/rank/2025/summary-by-size")
def get_rank_2025_summary_by_size(
    exclude_org_name: str = Query("삼성전자", description="합계에서 제외할 조직명 (정확히 일치)"),
    years: str = Query("2025,2026", description="콤마 구분 연도 리스트(예: 2025,2026)"),
) -> dict:
    try:
        years_list = [int(str(y).strip()) for y in (years.split(",") if years else []) if str(y).strip()]
        return db.get_rank_2025_summary_by_size(exclude_org_name=exclude_org_name, years=years_list)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/monthly-amounts/summary")
def get_performance_monthly_amounts_summary(
    from_month: str = Query("2025-01", description="시작 YYYY-MM"),
    to_month: str = Query("2026-12", description="종료 YYYY-MM"),
    team: str | None = Query(None, description="edu1|edu2 (선택)"),
) -> dict:
    try:
        return db.get_perf_monthly_amounts_summary(from_month=from_month, to_month=to_month, team=team)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/monthly-amounts/deals")
def get_performance_monthly_amounts_deals(
    segment: str = Query(..., description="세그먼트 키"),
    row: str = Query(..., description="CONTRACT|CONFIRMED|HIGH"),
    month: str = Query(..., description="YYMM (예: 2501)"),
    team: str | None = Query(None, description="edu1|edu2 (선택)"),
) -> dict:
    try:
        return db.get_perf_monthly_amounts_deals(segment=segment, row=row, month=month, team=team)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/monthly-inquiries/summary")
def get_performance_monthly_inquiries_summary(
    from_month: str = Query("2025-01", description="시작 YYYY-MM"),
    to_month: str = Query("2026-12", description="종료 YYYY-MM"),
    team: str | None = Query(None, description="edu1|edu2 (선택)"),
    debug: bool = Query(False, description="디버그/캐시우회 플래그"),
) -> dict:
    try:
        return db.get_perf_monthly_inquiries_summary(from_month=from_month, to_month=to_month, team=team, debug=debug)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/monthly-inquiries/deals")
def get_performance_monthly_inquiries_deals(
    segment: str = Query(..., description="세그먼트 키 (기업 규모)"),
    row: str = Query(..., description="과정포맷||카테고리그룹"),
    month: str = Query(..., description="YYMM (예: 2501)"),
    team: str | None = Query(None, description="edu1|edu2 (선택)"),
    debug: bool = Query(False, description="디버그/캐시우회 플래그"),
) -> dict:
    try:
        return db.get_perf_monthly_inquiries_deals(segment=segment, row=row, month=month, team=team, debug=debug)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/monthly-close-rate/summary")
def get_performance_monthly_close_rate_summary(
    from_month: str = Query("2025-01", alias="from", description="시작 YYYY-MM"),
    to_month: str = Query("2026-12", alias="to", description="종료 YYYY-MM"),
    cust: str = Query("all", description="all|new|existing"),
    scope: str = Query("all", description="all|corp_group|edu1|edu2|edu1_p1|edu1_p2|edu2_p1|edu2_p2|edu2_online"),
) -> dict:
    try:
        return db.get_perf_monthly_close_rate_summary(from_month=from_month, to_month=to_month, cust=cust, scope=scope)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/monthly-close-rate/deals")
def get_performance_monthly_close_rate_deals(
    segment: str = Query(..., description="세그먼트 키 (기업 규모)"),
    row: str | None = Query(None, description="course_group||metric"),
    month: str = Query(..., description="YYMM (예: 2501)"),
    cust: str = Query("all", description="all|new|existing"),
    scope: str = Query("all", description="all|corp_group|edu1|edu2|edu1_p1|edu1_p2|edu2_p1|edu2_p2|edu2_online"),
    course: str | None = Query(None, description="course_group (row 미제공 시 fallback)"),
    metric: str | None = Query(None, description="metric (row 미제공 시 fallback)"),
) -> dict:
    try:
        if not row and course and metric:
            row = f"{course}||{metric}"
        if not row:
            raise HTTPException(status_code=400, detail="row or course+metric is required")
        return db.get_perf_monthly_close_rate_deals(segment=segment, row=row, month=month, cust=cust, scope=scope)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/pl-progress-2026/summary")
def get_pl_progress_summary(year: int = Query(2026, description="연도 (기본 2026)")) -> dict:
    try:
        return db.get_pl_progress_summary(year=year)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/performance/pl-progress-2026/deals")
def get_pl_progress_deals(
    year: int = Query(2026, description="연도 (기본 2026)"),
    month: str = Query(..., description="YYMM (예: 2601)"),
    rail: str = Query(..., description="TOTAL|ONLINE|OFFLINE"),
    variant: str = Query("E", description="T|E (T는 드릴다운 없음)"),
    limit: int = Query(500, ge=1, le=2000),
    offset: int = Query(0, ge=0),
) -> dict:
    try:
        return db.get_pl_progress_deals(
            year=year,
            month=month,
            rail=rail,
            variant=variant,
            limit=limit,
            offset=offset,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/report/counterparty-risk")
def get_counterparty_risk_report(
    date: str | None = Query(None, description="YYYY-MM-DD (없으면 today)"),
    mode: str = Query("offline", description='리포트 모드 ("offline"|"online")'),
) -> dict:
    try:
        if date:
            # Validate date format early for clear 400
            datetime.fromisoformat(date)
        if mode not in {"offline", "online"}:
            raise HTTPException(status_code=400, detail="Invalid mode")
        try:
            return get_cached_report(as_of=date, mode=mode)
        except FileNotFoundError:
            run_daily_counterparty_risk_job(as_of_date=date, force=True, mode=mode)
            return get_cached_report(as_of=date, mode=mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {exc}")
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/report/counterparty-risk/recompute")
def recompute_counterparty_risk_report(
    date: str | None = Query(None, description="YYYY-MM-DD (없으면 today)"),
    mode: str = Query("offline", description='리포트 모드 ("offline"|"online")'),
) -> dict:
    try:
        if date:
            datetime.fromisoformat(date)
        if mode not in {"offline", "online"}:
            raise HTTPException(status_code=400, detail="Invalid mode")
        return run_daily_counterparty_risk_job(as_of_date=date, force=True, mode=mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {exc}")
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/report/counterparty-risk/status")
def get_counterparty_risk_status(
    mode: str | None = Query(None, description='리포트 모드 ("offline"|"online"), 없으면 전체 반환'),
) -> dict:
    if mode is None:
        return {
            "offline": _load_status(mode="offline"),
            "online": _load_status(mode="online"),
            "modes_available": ["offline", "online"],
        }
    if mode not in {"offline", "online"}:
        raise HTTPException(status_code=400, detail="Invalid mode")
    return _load_status(mode=mode)


def _parse_bool(val: bool | str | None, default: bool = False) -> bool:
    if isinstance(val, bool):
        return val
    if val is None:
        return default
    s = str(val).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


@router.post("/llm/target-attainment")
def post_target_attainment(
    req: TargetAttainmentRequest,
    debug: bool = Query(False, description="attach __meta when true"),
    nocache: bool = Query(False, description="skip cache when true"),
    include_input: bool = Query(False, description="include __llm_input when true"),
) -> dict:
    try:
        payload_dict = req.model_dump()
        try:
            size = validate_payload_limits(payload_dict)
        except ValueError:
            raise HTTPException(
                status_code=413,
                detail={
                    "error": "PAYLOAD_TOO_LARGE",
                    "max_bytes": MAX_TARGET_ATTAINMENT_REQUEST_BYTES,
                    "bytes": len(json.dumps(payload_dict, ensure_ascii=False).encode("utf-8")),
                    "hint": "upperOrg 1개 그룹만 포함되도록 compact JSON을 축소하세요.",
                },
            )
        return run_target_attainment(
            req,
            debug=debug,
            payload_bytes=size,
            nocache=_parse_bool(nocache),
            include_input=_parse_bool(include_input),
        )
    except HTTPException:
        raise
    except Exception as exc:  # pragma: no cover - defensive
        return {"error": "TARGET_ATTAINMENT_INTERNAL_ERROR", "message": str(exc)}


@router.post("/llm/daily-report-v2/pipeline")
def post_daily_report_v2_pipeline(
    payload: dict,
    pipeline_id: str = Query(..., description='파이프라인 ID (예: "daily.part_rollup", "row.target_attainment")'),
    variant: str = Query("offline", description='모드 ("offline"|"online")'),
    debug: bool = Query(False, description="attach __meta when true"),
    nocache: bool = Query(False, description="skip cache when true"),
) -> dict:
    try:
        return run_daily_report_v2_pipeline(pipeline_id, payload, variant=variant, debug=debug, nocache=_parse_bool(nocache))
    except Exception as exc:  # pragma: no cover - defensive
        return {"error": "DAILY_REPORT_V2_PIPELINE_ERROR", "message": str(exc)}


@router.get("/rank/2025-deals-people")
def get_rank_2025_deals_people(
    size: str = Query("대기업", description='조직 규모 필터 (예: "대기업", "전체")')
) -> dict:
    try:
        return {"items": db.get_rank_2025_deals_people(size=size)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))

@router.get("/rank/2025-top100-counterparty-dri")
def get_rank_2025_top100_counterparty_dri(
    size: str = Query("대기업", description='조직 규모 필터 (예: "대기업", "전체")'),
    limit: int | None = Query(
        None,
        ge=1,
        le=200_000,
        description="최대 반환 수 (미지정 시 전체 반환)",
    ),
    offset: int = Query(0, ge=0, description="org 목록 offset (limit 단위, limit 미지정 시 무시)"),
    debug: bool = Query(False, description="override 매칭 진단 포함 여부"),
) -> dict:
    try:
        return db.get_rank_2025_top100_counterparty_dri(size=size, limit=limit, offset=offset, debug=debug)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))

@router.get("/rank/2025-counterparty-dri/detail")
def get_rank_counterparty_dri_detail(orgId: str = Query(...), upperOrg: str = Query(...)) -> dict:
    try:
        return db.get_rank_2025_counterparty_detail(org_id=orgId, upper_org=upperOrg)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/rank/2025-top100-counterparty-dri/targets-summary")
def get_rank_counterparty_dri_targets_summary(
    size: str = Query("대기업", description='조직 규모 필터 (예: "대기업", "전체")')
) -> dict:
    try:
        return db.get_rank_2025_counterparty_dri_targets_summary(size=size)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/rank/won-industry-summary")
def get_rank_won_industry_summary(
    size: str = Query("전체", description='조직 규모 필터 (예: "대기업", "중견기업", "전체")')
) -> dict:
    try:
        return {"items": db.get_won_industry_summary(size=size)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}/won-summary")
def get_won_summary(org_id: str) -> dict:
    try:
        return {"items": db.get_won_summary_by_upper_org(org_id=org_id)}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))

@router.get("/orgs/{org_id}/won-groups-json")
def get_won_groups_json(org_id: str) -> dict:
    try:
        return db.get_won_groups_json(org_id=org_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}/won-groups-json-compact")
def get_won_groups_json_compact(org_id: str) -> dict:
    try:
        raw = db.get_won_groups_json(org_id=org_id)
        return compact_won_groups_json(raw)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}/won-groups-markdown-compact")
def get_won_groups_markdown_compact(
    org_id: str,
    upper_org: str | None = Query(None, description="상위 조직 필터"),
    max_deals: int = Query(200, ge=1, le=500),
    max_people: int = Query(60, ge=1, le=500),
    deal_memo_limit: int = Query(10, ge=1, le=50),
    memo_max_chars: int = Query(240, ge=50, le=500),
    redact_phone: bool = Query(True),
    max_output_chars: int = Query(200_000, ge=10_000, le=1_000_000),
    format: str = Query("text", regex="^(text|json)$"),
) -> Any:
    try:
        uppers = [upper_org] if upper_org else None
        raw = db.get_won_groups_json(org_id=org_id, target_uppers=uppers)
        compact = compact_won_groups_json(raw)
        md = won_groups_compact_to_markdown(
            compact,
            scope_label="UPPER_SELECTED" if upper_org else "ORG_ALL",
            max_people=max_people,
            max_deals=max_deals,
            deal_memo_limit=deal_memo_limit,
            memo_max_chars=memo_max_chars,
            redact_phone=redact_phone,
            max_output_chars=max_output_chars,
        )
        if format == "json":
            return {"schema_version": "won-groups-json/compact-md-v1.1", "markdown": md}
        return PlainTextResponse(md, media_type="text/plain; charset=utf-8")
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}/statepath")
def get_statepath(org_id: str) -> dict:
    try:
        raw = db.get_won_groups_json(org_id=org_id)
        compact = compact_won_groups_json(raw)
        item = build_statepath(compact)
        return {"item": item}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/statepath/portfolio-2425")
def get_statepath_portfolio(
    segment: str = Query("전체", description="대기업/중견기업/중소기업/공공기관/대학교/기타/미입력"),
    legacySizeGroup: str | None = Query(None, alias="sizeGroup"),
    search: str | None = Query(None, description="조직명 검색"),
    sort: str = Query("won2025_desc"),
    limit: int = Query(500, ge=1, le=2000),
    offset: int = Query(0, ge=0),
    riskOnly: bool = False,
    hasOpen: bool = False,
    hasScaleUp: bool = False,
    companyDir: str = Query("all"),
    seed: str = Query("all"),
    rail: str = Query("all"),
    railDir: str = Query("all"),
    companyFrom: str = Query("all"),
    companyTo: str = Query("all"),
    cell: str = Query("all"),
    cellEvent: str = Query("all"),
) -> dict:
    try:
        filters = {
            "riskOnly": riskOnly,
            "hasOpen": hasOpen,
            "hasScaleUp": hasScaleUp,
            "companyDir": companyDir,
            "seed": seed,
            "rail": rail,
            "railDir": railDir,
            "companyFrom": companyFrom,
            "companyTo": companyTo,
            "cell": cell,
            "cellEvent": cellEvent,
        }
        chosen_segment = legacySizeGroup or segment
        return db.get_statepath_portfolio(
            size_group=chosen_segment,
            search=search,
            filters=filters,
            sort=sort,
            limit=limit,
            offset=offset,
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}/statepath-2425")
def get_statepath_detail(org_id: str) -> dict:
    try:
        item = db.get_statepath_detail(org_id)
        if not item:
            raise HTTPException(status_code=404, detail="Organization not found")
        return {"item": item}
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/orgs/{org_id}")
def get_org(org_id: str) -> dict:
    try:
        match = db.get_org_by_id(org_id)
        if not match:
            raise HTTPException(status_code=404, detail="Organization not found")
        return {"item": match}
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
