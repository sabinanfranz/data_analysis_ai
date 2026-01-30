#!/usr/bin/env python3
"""
Ad-hoc KPI + Audit XLSX Export (Won 고객사/GenAI 비중)

Reads salesmap_latest.db in read-only mode, computes 3 KPIs, writes
stdout + JSON summary, and emits an audit-friendly XLSX with per-KPI
denominator/numerator datasets.
"""

import argparse
import datetime
import json
import os
import sqlite3
import sys
import urllib.parse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError as exc:  # pragma: no cover - helpful runtime guard
    sys.stderr.write(
        "Missing dependency: openpyxl. Install with `pip install -r requirements.txt` "
        "or `pip install openpyxl>=3.1.0`.\n"
    )
    raise SystemExit(1) from exc


# --- SQL SSOT ---
COMMON_WITH = """
WITH base AS (
  SELECT
    rowid                                   AS deal_rowid,
    NULLIF(TRIM(organizationId), '')         AS org_id,
    COALESCE(TRIM("상태"), '')               AS status_raw,
    COALESCE(TRIM("계약 체결일"), '')        AS contract_date,
    COALESCE(TRIM("과정포맷"), '')           AS course_format,
    COALESCE(TRIM("카테고리"), '')           AS category,
    COALESCE(TRIM("금액"), '')               AS amount_raw,
    COALESCE(TRIM("수강시작일"), '')         AS start_raw,
    COALESCE(TRIM("수강종료일"), '')         AS end_raw
  FROM deal
),
typed0 AS (
  SELECT
    deal_rowid,
    org_id,
    status_raw,
    amount_raw,
    LOWER(status_raw) AS status_lc,
    CASE
      WHEN contract_date <> '' AND contract_date GLOB '[0-9][0-9][0-9][0-9]*'
        THEN CAST(substr(contract_date, 1, 4) AS INTEGER)
      ELSE NULL
    END AS contract_year,
    contract_date,
    course_format,
    category,
    CASE
      WHEN course_format IN ('구독제(온라인)', '선택구매(온라인)', '포팅')
           OR course_format LIKE '포팅%'
        THEN 1 ELSE 0
    END AS is_excluded_online_format,
    CASE
      WHEN amount_raw = '' THEN NULL
      ELSE
        CASE
          WHEN REPLACE(REPLACE(REPLACE(REPLACE(amount_raw, ',', ''), '원', ''), ' ', ''), '\t', '') <> ''
           AND REPLACE(REPLACE(REPLACE(REPLACE(amount_raw, ',', ''), '원', ''), ' ', ''), '\t', '') NOT GLOB '*[^0-9.]*'
            THEN CAST(REPLACE(REPLACE(REPLACE(REPLACE(amount_raw, ',', ''), '원', ''), ' ', ''), '\t', '') AS REAL)
          ELSE NULL
        END
    END AS amount_num,
    start_raw,
    end_raw,
    CASE
      WHEN start_raw = '' THEN NULL
      WHEN start_raw GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]*'
        THEN substr(start_raw, 1, 10)
      WHEN start_raw GLOB '[0-9][0-9][0-9][0-9]/[0-9][0-9]/[0-9][0-9]*'
        THEN REPLACE(substr(start_raw, 1, 10), '/', '-')
      WHEN start_raw GLOB '[0-9][0-9][0-9][0-9].[0-9][0-9].[0-9][0-9]*'
        THEN REPLACE(substr(start_raw, 1, 10), '.', '-')
      WHEN start_raw GLOB '[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]*'
        THEN substr(start_raw, 1, 4) || '-' || substr(start_raw, 5, 2) || '-' || substr(start_raw, 7, 2)
      ELSE NULL
    END AS start_date,
    CASE
      WHEN end_raw = '' THEN NULL
      WHEN end_raw GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]*'
        THEN substr(end_raw, 1, 10)
      WHEN end_raw GLOB '[0-9][0-9][0-9][0-9]/[0-9][0-9]/[0-9][0-9]*'
        THEN REPLACE(substr(end_raw, 1, 10), '/', '-')
      WHEN end_raw GLOB '[0-9][0-9][0-9][0-9].[0-9][0-9].[0-9][0-9]*'
        THEN REPLACE(substr(end_raw, 1, 10), '.', '-')
      WHEN end_raw GLOB '[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]*'
        THEN substr(end_raw, 1, 4) || '-' || substr(end_raw, 5, 2) || '-' || substr(end_raw, 7, 2)
      ELSE NULL
    END AS end_date
  FROM base
  WHERE org_id IS NOT NULL
),
typed AS (
  SELECT
    typed0.*,
    CASE
      WHEN start_date IS NOT NULL AND end_date IS NOT NULL
        THEN CAST(julianday(end_date) - julianday(start_date) AS INTEGER) + 1
      ELSE NULL
    END AS duration_days
  FROM typed0
)
""".strip()  # <-- 세미콜론 금지!

GENAI_CATEGORY = "생성형AI"


def build_sql(select_sql: str, extra_ctes: list[str] | None = None) -> str:
    """
    extra_ctes: ["flags AS (...)" , ...] (WITHOUT leading WITH).
    """
    sql = COMMON_WITH
    if extra_ctes:
        sql += "\n, " + "\n, ".join(extra_ctes)
    sql += "\n" + select_sql.strip().rstrip(";")
    # Guard: prevent accidental second WITH
    with_count = (1 if sql.startswith("WITH ") else 0) + sql.count("\nWITH ")
    if with_count > 1:
        raise ValueError("SQL has multiple WITH clauses. Use extra_ctes with commas, not another WITH.")
    return sql


QUERY_KPI1_WON_EVER = build_sql(
    """
SELECT COUNT(DISTINCT org_id) AS orgs_with_won_ever
FROM typed
WHERE status_lc='won'
"""
)

QUERY_KPI1_WON_2025_NEW = build_sql(
    select_sql="""
SELECT COUNT(*) AS orgs_won_2025_but_not_2024
FROM flags
WHERE won_2024=0 AND won_2025=1
""",
    extra_ctes=[
        """
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 THEN 1 ELSE 0 END) AS won_2025
  FROM typed
  GROUP BY org_id
)
"""
    ],
)

QUERY_KPI2 = build_sql(
    select_sql="""
SELECT
  COUNT(*) AS deal_count_total,
  COUNT(amount_num) AS deal_count_amount_ok,
  (COUNT(*) - COUNT(amount_num)) AS deal_count_amount_missing,
  SUM(CASE WHEN category='""" + GENAI_CATEGORY + """' THEN amount_num ELSE 0 END) AS genai_amount_sum,
  SUM(amount_num) AS total_amount_sum,
  CASE
    WHEN SUM(amount_num) > 0
      THEN ROUND(SUM(CASE WHEN category='""" + GENAI_CATEGORY + """' THEN amount_num ELSE 0 END) * 1.0 / SUM(amount_num), 6)
    ELSE NULL
  END AS genai_share
FROM scope
""",
    extra_ctes=[
        """
scope AS (
  SELECT *
  FROM typed
  WHERE status_lc='won'
    AND contract_year=2025
    AND is_excluded_online_format=0
)
"""
    ],
)

QUERY_KPI3 = build_sql(
    select_sql="""
SELECT
  (SELECT COUNT(*) FROM denom) AS denom_orgs_2024_nonai,
  (SELECT COUNT(*) FROM denom WHERE ai_2025=1) AS num_orgs_new_ai_2025,
  CASE
    WHEN (SELECT COUNT(*) FROM denom) > 0
      THEN ROUND((SELECT COUNT(*) FROM denom WHERE ai_2025=1) * 1.0 / (SELECT COUNT(*) FROM denom), 6)
    ELSE NULL
  END AS ratio_new_ai_2025
""",
    extra_ctes=[
        """
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2025
  FROM typed
  GROUP BY org_id
),
denom AS (
  SELECT *
  FROM flags
  WHERE won_2024=1 AND ai_2024=0
)
"""
    ],
)

QUERY_WON_EVER_DEALS = build_sql(
    """
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  status_raw, status_lc, category, course_format, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
"""
)

QUERY_WON_EVER_ORGS = build_sql(
    select_sql="""
SELECT
  org_id,
  COUNT(*) AS won_deal_count,
  MIN(contract_year) AS first_won_year,
  MAX(contract_year) AS last_won_year
FROM won
GROUP BY org_id
ORDER BY won_deal_count DESC
""",
    extra_ctes=[
        """
won AS (SELECT * FROM typed WHERE status_lc='won')
"""
    ],
)

QUERY_NEW_2025_ORGS = build_sql(
    select_sql="""
SELECT
  org_id,
  0 AS won_2024,
  1 AS won_2025,
  COUNT(*) AS won_2025_deal_count,
  MIN(contract_date) AS first_won_2025_date
FROM won_2025
GROUP BY org_id
""",
    extra_ctes=[
        """
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 THEN 1 ELSE 0 END) AS won_2025
  FROM typed
  GROUP BY org_id
),
new_org AS (SELECT org_id FROM flags WHERE won_2024=0 AND won_2025=1),
won_2025 AS (
  SELECT * FROM typed
  WHERE status_lc='won' AND contract_year=2025
    AND org_id IN (SELECT org_id FROM new_org)
)
"""
    ],
)

QUERY_NEW_2025_DEALS = build_sql(
    select_sql="""
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  category, course_format, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2025
  AND org_id IN (SELECT org_id FROM new_org)
""",
    extra_ctes=[
        """
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 THEN 1 ELSE 0 END) AS won_2025
  FROM typed GROUP BY org_id
),
new_org AS (
  SELECT org_id FROM flags WHERE won_2024=0 AND won_2025=1
)
"""
    ],
)

QUERY_KPI2_SCOPE = build_sql(
    """
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  course_format, is_excluded_online_format,
  category, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2025
  AND is_excluded_online_format=0
"""
)

QUERY_KPI2_GENAI = build_sql(
    """
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  course_format, category, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2025
  AND is_excluded_online_format=0
  AND category='""" + GENAI_CATEGORY + """'
"""
)

QUERY_KPI2_DROPPED_NULL = build_sql(
    """
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  course_format, category, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2025
  AND is_excluded_online_format=0
  AND amount_num IS NULL
"""
)

QUERY_KPI3_DENOM_ORGS = build_sql(
    select_sql="""
SELECT *
FROM flags
WHERE won_2024=1 AND ai_2024=0
""",
    extra_ctes=[
        """
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2025
  FROM typed GROUP BY org_id
)
"""
    ],
)

QUERY_KPI3_DENOM_DEALS = build_sql(
    select_sql="""
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  category, course_format, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2024
  AND org_id IN (SELECT org_id FROM denom)
""",
    extra_ctes=[
        """
denom AS (
  SELECT org_id
  FROM (
    SELECT
      org_id,
      MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
      MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2024
    FROM typed GROUP BY org_id
  )
  WHERE won_2024=1 AND ai_2024=0
)
"""
    ],
)

QUERY_KPI3_NUM_ORGS = build_sql(
    select_sql="""
SELECT *
FROM denom
WHERE ai_2025=1
""",
    extra_ctes=[
        """
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2025
  FROM typed GROUP BY org_id
),
denom AS (
  SELECT * FROM flags WHERE won_2024=1 AND ai_2024=0
)
"""
    ],
)

QUERY_KPI3_NUM_DEALS = build_sql(
    select_sql="""
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  category, course_format, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2025
  AND category='""" + GENAI_CATEGORY + """'
  AND org_id IN (SELECT org_id FROM denom)
""",
    extra_ctes=[
        """
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 THEN 1 ELSE 0 END) AS won_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2024,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND category='""" + GENAI_CATEGORY + """' THEN 1 ELSE 0 END) AS ai_2025
  FROM typed GROUP BY org_id
),
denom AS (
  SELECT org_id FROM flags WHERE won_2024=1 AND ai_2024=0 AND ai_2025=1
)
"""
    ],
)

QUERY_KPI4_SUMMARY = build_sql(
    select_sql="""
SELECT
  (SELECT COUNT(*) FROM denom) AS denom_orgs_2024_nonai_exclfmt,
  (SELECT COUNT(*) FROM denom WHERE ai_2025_exclfmt=1) AS num_orgs_new_ai_2025_exclfmt,
  CASE
    WHEN (SELECT COUNT(*) FROM denom) > 0
      THEN ROUND(
        (SELECT COUNT(*) FROM denom WHERE ai_2025_exclfmt=1) * 1.0
        / (SELECT COUNT(*) FROM denom)
      , 6)
    ELSE NULL
  END AS ratio_new_ai_2025_exclfmt
""",
    extra_ctes=[
        f"""
flags AS (
  SELECT
    org_id,
    MAX(CASE
          WHEN status_lc='won'
           AND contract_year=2024
           AND is_excluded_online_format=0
          THEN 1 ELSE 0 END
    ) AS won_2024_exclfmt,

    MAX(CASE
          WHEN status_lc='won'
           AND contract_year=2024
           AND is_excluded_online_format=0
           AND category='{GENAI_CATEGORY}'
          THEN 1 ELSE 0 END
    ) AS ai_2024_exclfmt,

    MAX(CASE
          WHEN status_lc='won'
           AND contract_year=2025
           AND is_excluded_online_format=0
           AND category='{GENAI_CATEGORY}'
          THEN 1 ELSE 0 END
    ) AS ai_2025_exclfmt
  FROM typed
  GROUP BY org_id
),
denom AS (
  SELECT *
  FROM flags
  WHERE won_2024_exclfmt=1
    AND ai_2024_exclfmt=0
)
""".strip()
    ],
)

QUERY_KPI4_DENOM_ORGS = build_sql(
    select_sql="""
SELECT
  org_id,
  won_2024_exclfmt,
  ai_2024_exclfmt,
  ai_2025_exclfmt
FROM denom
ORDER BY ai_2025_exclfmt DESC, org_id
""",
    extra_ctes=[
        f"""
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 THEN 1 ELSE 0 END) AS won_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2025_exclfmt
  FROM typed GROUP BY org_id
),
denom AS (
  SELECT * FROM flags
  WHERE won_2024_exclfmt=1 AND ai_2024_exclfmt=0
)
""".strip()
    ],
)

QUERY_KPI4_DENOM_DEALS_2024 = build_sql(
    select_sql="""
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  course_format, is_excluded_online_format,
  category, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2024
  AND is_excluded_online_format=0
  AND org_id IN (SELECT org_id FROM denom)
ORDER BY org_id, contract_date, deal_rowid
""",
    extra_ctes=[
        f"""
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 THEN 1 ELSE 0 END) AS won_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2025_exclfmt
  FROM typed GROUP BY org_id
),
denom AS (
  SELECT * FROM flags
  WHERE won_2024_exclfmt=1 AND ai_2024_exclfmt=0
)
""".strip()
    ],
)

QUERY_KPI4_NUM_ORGS = build_sql(
    select_sql="""
SELECT
  org_id,
  won_2024_exclfmt,
  ai_2024_exclfmt,
  ai_2025_exclfmt
FROM denom
WHERE ai_2025_exclfmt=1
ORDER BY org_id
""",
    extra_ctes=[
        f"""
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 THEN 1 ELSE 0 END) AS won_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2025_exclfmt
  FROM typed GROUP BY org_id
),
denom AS (
  SELECT * FROM flags
  WHERE won_2024_exclfmt=1 AND ai_2024_exclfmt=0
)
""".strip()
    ],
)

QUERY_KPI4_NUM_DEALS_2025_AI = build_sql(
    select_sql=f"""
SELECT
  deal_rowid, org_id, contract_date, contract_year,
  course_format, is_excluded_online_format,
  category, amount_raw, amount_num
FROM typed
WHERE status_lc='won'
  AND contract_year=2025
  AND is_excluded_online_format=0
  AND category='{GENAI_CATEGORY}'
  AND org_id IN (SELECT org_id FROM denom WHERE ai_2025_exclfmt=1)
ORDER BY org_id, contract_date, deal_rowid
""",
    extra_ctes=[
        f"""
flags AS (
  SELECT
    org_id,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 THEN 1 ELSE 0 END) AS won_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2024 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2024_exclfmt,
    MAX(CASE WHEN status_lc='won' AND contract_year=2025 AND is_excluded_online_format=0 AND category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS ai_2025_exclfmt
  FROM typed GROUP BY org_id
),
denom AS (
  SELECT * FROM flags
  WHERE won_2024_exclfmt=1 AND ai_2024_exclfmt=0
)
""".strip()
    ],
)

QUERY_KPI5_COUNT_SHARE_2025_EXCLFMT = build_sql(
    f"""
SELECT
  COUNT(*) AS won_deal_count_total_2025_exclfmt,
  SUM(CASE WHEN category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) AS won_deal_count_genai_2025_exclfmt,
  CASE
    WHEN COUNT(*) > 0
      THEN ROUND(
        SUM(CASE WHEN category='{GENAI_CATEGORY}' THEN 1 ELSE 0 END) * 1.0 / COUNT(*),
      6)
    ELSE NULL
  END AS genai_deal_share_2025_exclfmt
FROM typed
WHERE status_lc='won'
  AND contract_year=2025
  AND is_excluded_online_format=0
"""
)

QUERY_KPI6_DURATION_BUCKET_COUNTS_2025_GENAI_EXCLFMT = build_sql(
    select_sql="""
SELECT
  (SELECT COUNT(*) FROM scope) AS scope_deal_count,
  (SELECT COUNT(*) FROM valid) AS valid_deal_count,
  (SELECT COUNT(*) FROM scope) - (SELECT COUNT(*) FROM valid) AS dropped_missing_or_invalid_dates,

  SUM(CASE WHEN duration_days = 1 THEN 1 ELSE 0 END) AS cnt_1day,
  SUM(CASE WHEN duration_days BETWEEN 2  AND 29  THEN 1 ELSE 0 END) AS cnt_gt1_le1m,
  SUM(CASE WHEN duration_days BETWEEN 30 AND 89  THEN 1 ELSE 0 END) AS cnt_ge1m_le3m,
  SUM(CASE WHEN duration_days BETWEEN 90 AND 179 THEN 1 ELSE 0 END) AS cnt_ge3m_le6m,
  SUM(CASE WHEN duration_days BETWEEN 180 AND 365 THEN 1 ELSE 0 END) AS cnt_ge6m_le12m,
  SUM(CASE WHEN duration_days >= 366 THEN 1 ELSE 0 END) AS cnt_gt12m
FROM valid
""",
    extra_ctes=[f"""
scope AS (
  SELECT *
  FROM typed
  WHERE status_lc='won'
    AND contract_year=2025
    AND is_excluded_online_format=0
    AND category='{GENAI_CATEGORY}'
),
valid AS (
  SELECT *
  FROM scope
  WHERE start_date IS NOT NULL
    AND end_date IS NOT NULL
    AND duration_days IS NOT NULL
    AND duration_days >= 1
    AND julianday(end_date) >= julianday(start_date)
)
""".strip()],
)

QUERY_KPI6_DURATION_VALID_DEALS_2025_GENAI_EXCLFMT = build_sql(
    select_sql="""
SELECT
  deal_rowid,
  org_id,
  contract_date,
  course_format,
  category,
  start_raw,
  end_raw,
  start_date,
  end_date,
  duration_days,
  CASE
    WHEN duration_days = 1 THEN 'B01_1day'
    WHEN duration_days BETWEEN 2 AND 29 THEN 'B02_gt1_le1m'
    WHEN duration_days BETWEEN 30 AND 89 THEN 'B03_ge1m_le3m'
    WHEN duration_days BETWEEN 90 AND 179 THEN 'B04_ge3m_le6m'
    WHEN duration_days BETWEEN 180 AND 365 THEN 'B05_ge6m_le12m'
    WHEN duration_days >= 366 THEN 'B06_gt12m'
    ELSE 'B00_unknown'
  END AS duration_bucket
FROM valid
ORDER BY duration_days ASC, deal_rowid ASC
""",
    extra_ctes=[f"""
scope AS (
  SELECT *
  FROM typed
  WHERE status_lc='won'
    AND contract_year=2025
    AND is_excluded_online_format=0
    AND category='{GENAI_CATEGORY}'
),
valid AS (
  SELECT *
  FROM scope
  WHERE start_date IS NOT NULL
    AND end_date IS NOT NULL
    AND duration_days IS NOT NULL
    AND duration_days >= 1
    AND julianday(end_date) >= julianday(start_date)
)
""".strip()],
)

QUERY_KPI6_DURATION_DROPPED_DEALS_2025_GENAI_EXCLFMT = build_sql(
    select_sql="""
SELECT
  deal_rowid,
  org_id,
  contract_date,
  course_format,
  category,
  start_raw,
  end_raw,
  start_date,
  end_date,
  duration_days,
  CASE
    WHEN start_date IS NULL OR end_date IS NULL THEN 'MISSING_OR_UNPARSABLE_DATE'
    WHEN duration_days IS NULL THEN 'DURATION_NULL'
    WHEN duration_days < 1 THEN 'DURATION_LT_1'
    WHEN julianday(end_date) < julianday(start_date) THEN 'END_BEFORE_START'
    ELSE 'OTHER'
  END AS drop_reason
FROM scope
WHERE NOT (
  start_date IS NOT NULL
  AND end_date IS NOT NULL
  AND duration_days IS NOT NULL
  AND duration_days >= 1
  AND julianday(end_date) >= julianday(start_date)
)
ORDER BY drop_reason, deal_rowid
""",
    extra_ctes=[f"""
scope AS (
  SELECT *
  FROM typed
  WHERE status_lc='won'
    AND contract_year=2025
    AND is_excluded_online_format=0
    AND category='{GENAI_CATEGORY}'
)
""".strip()],
)


REQUIRED_DEAL_COLS = {
    "상태",
    "계약 체결일",
    "금액",
    "과정포맷",
    "카테고리",
    "organizationId",
    "수강시작일",
    "수강종료일",
}

PERCENT_COLUMNS = {
    "genai_share",
    "ratio_new_ai_2025",
    "ratio_new_ai_2025_exclfmt",
    "genai_deal_share_2025_exclfmt",
    # duration bucket share not computed as percent, keep for potential additions
    "value_pct",
}
AMOUNT_COLUMNS = {"amount_num", "genai_amount_sum", "total_amount_sum"}


def connect_readonly(db_path: str) -> sqlite3.Connection:
    abs_path = os.path.abspath(db_path)
    uri_path = urllib.parse.quote(os.path.abspath(abs_path))
    uri = f"file:{uri_path}?mode=ro"
    return sqlite3.connect(uri, uri=True)


def run_query(conn: sqlite3.Connection, sql: str, params=None):
    sql = sql.strip().rstrip(";")
    try:
        cur = conn.execute(sql, params or [])
        cols = [d[0] for d in cur.description] if cur.description else []
        rows = cur.fetchall()
        return cols, rows
    except Exception as e:  # pragma: no cover - defensive logging
        print("SQLite error:", e)
        print("----- SQL BEGIN -----")
        print(sql)
        print("----- SQL END -----")
        raise


def rows_to_dicts(columns, rows):
    return [dict(zip(columns, row)) for row in rows]


def get_org_name_map(conn: sqlite3.Connection):
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='organization'"
    )
    if cur.fetchone() is None:
        return {}

    info = conn.execute("PRAGMA table_info(organization)").fetchall()
    col_names = [row[1] for row in info]

    id_candidates = ["id", "organizationId", "org_id"]
    name_candidates = ["name", "organizationName", "기업명", "org_name"]

    id_col = next((c for c in id_candidates if c in col_names), None)
    name_col = next((c for c in name_candidates if c in col_names), None)
    if not id_col or not name_col:
        return {}

    mapping = {}
    for row in conn.execute(f'SELECT {id_col}, {name_col} FROM organization'):
        org_id, org_name = row
        if org_id is None:
            continue
        mapping[str(org_id).strip()] = (org_name or "").strip()
    return mapping


def append_org_names(dict_rows, org_map):
    if not org_map:
        return dict_rows
    for row in dict_rows:
        org_id = row.get("org_id")
        row["org_name"] = org_map.get(org_id, "") if org_id else ""
    return dict_rows


def ensure_required_columns(conn: sqlite3.Connection):
    cols = {row[1] for row in conn.execute("PRAGMA table_info(deal)")}
    missing = REQUIRED_DEAL_COLS - cols
    if missing:
        raise SystemExit(
            f"Missing required columns in deal table: {', '.join(sorted(missing))}"
        )


def write_sheet(ws, dict_rows, header_order=None):
    rows = dict_rows or []
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())

    if header_order:
        headers = [h for h in header_order if h in all_keys or not rows]
        extra = [k for k in all_keys if k not in header_order]
        headers.extend(sorted(extra))
    else:
        headers = sorted(all_keys)

    # Ensure at least headers
    if not headers:
        headers = []
        if rows:
            headers = sorted(rows[0].keys())

    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for row in rows:
        ws.append([row.get(h, None) for h in headers])

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    # Number formats
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell, header in zip(row, headers):
            if header in PERCENT_COLUMNS and cell.value is not None:
                cell.number_format = "0.00%"
            elif header in AMOUNT_COLUMNS and cell.value is not None:
                cell.number_format = "#,##0.00"
            elif header.lower().endswith("_count") and cell.value is not None:
                cell.number_format = "#,##0"


def build_readme_sheet_data(db_path: str, generated_at: str):
    return [
        {
            "item": "db_path",
            "value": os.path.abspath(db_path),
        },
        {"item": "generated_at_utc", "value": generated_at},
        {"item": "kpi1_1", "value": "누적 Won 고객사 수 (status=Won ever)"},
        {
            "item": "kpi1_2",
            "value": "2025 신규 계약 고객사 수 (2024 Won 없음 AND 2025 Won 있음)",
        },
        {
            "item": "kpi2",
            "value": "2025 Won 비온라인 딜 중 생성형AI 금액 비중 (금액 합 기준)",
        },
        {
            "item": "kpi3",
            "value": "2024 비AI 수강 고객사 대비 2025 신규 AI 수강 비율",
        },
        {
            "item": "kpi4",
            "value": "2024 비AI 수강 고객사 대비 2025 신규 AI 수강 비율 (구독제/선택구매/포팅 포맷 제외)",
        },
        {
            "item": "kpi5",
            "value": "2025 Won 비온라인(포맷 제외) 딜 중 생성형AI 딜 개수 비중 (분모/분자 시트: 20_KPI2_SCOPE_DEALS_2025, 21_KPI2_GENAI_DEALS_2025)",
        },
        {
            "item": "kpi6",
            "value": "2025 Won·포맷 제외·생성형AI 딜의 수강기간 버킷(1/2-30/31-90/91-180/181-365/366+) 개수 분포. duration_days=(종료-시작)+1, 시작/종료/순서 이상이면 dropped 시트(61) 참조",
        },
        {
            "item": "kpi6_buckets",
            "value": "버킷 경계: 1일, 2-29일, 30-89일, 90-179일, 180-365일, 366일 이상. duration_days=(end-start)+1.",
        },
        {
            "item": "online_excluded_formats",
            "value": "구독제(온라인), 선택구매(온라인), 포팅(+포팅 변형 포함)",
        },
        {
            "item": "amount_parse",
            "value": "금액은 콤마/원/공백 제거 후 숫자+소수점만 허용, 실패 시 NULL",
        },
        {
            "item": "contract_year_rule",
            "value": '계약 체결일 앞 4자리 숫자 → INTEGER, 없으면 NULL',
        },
        {
            "item": "kpi2_amount_null_note",
            "value": "amount_num NULL은 분모 건수에 포함되지만 합계 제외; 시트 22_KPI2_DROPPED_AMOUNT_NULL 참고",
        },
    ]


def main():
    parser = argparse.ArgumentParser(
        description="Ad-hoc KPI + Audit XLSX Export (Won 고객사/GenAI 비중)"
    )
    parser.add_argument("--db-path", default="salesmap_latest.db")
    parser.add_argument("--out-json", default="output/kpi_2024_2025_ai.json")
    parser.add_argument("--out-xlsx", default="output/kpi_2024_2025_ai_audit.xlsx")
    args = parser.parse_args()

    _required = [
        "QUERY_KPI1_WON_EVER",
        "QUERY_KPI1_WON_2025_NEW",
        "QUERY_KPI2",
        "QUERY_KPI3",
        "QUERY_KPI4_SUMMARY",
        "QUERY_KPI5_COUNT_SHARE_2025_EXCLFMT",
        "QUERY_KPI6_DURATION_BUCKET_COUNTS_2025_GENAI_EXCLFMT",
        "QUERY_KPI6_DURATION_VALID_DEALS_2025_GENAI_EXCLFMT",
        "QUERY_KPI6_DURATION_DROPPED_DEALS_2025_GENAI_EXCLFMT",
    ]
    missing = [k for k in _required if k not in globals()]
    if missing:
        raise RuntimeError(f"Missing query constants: {missing}")

    os.makedirs(os.path.dirname(os.path.abspath(args.out_json)), exist_ok=True)
    os.makedirs(os.path.dirname(os.path.abspath(args.out_xlsx)), exist_ok=True)

    # Connect read-only
    conn = connect_readonly(args.db_path)
    try:
        ensure_required_columns(conn)

        org_map = get_org_name_map(conn)

        # KPI calculations
        kpi1_1 = rows_to_dicts(*run_query(conn, QUERY_KPI1_WON_EVER))[0][
            "orgs_with_won_ever"
        ]
        kpi1_2 = rows_to_dicts(*run_query(conn, QUERY_KPI1_WON_2025_NEW))[0][
            "orgs_won_2025_but_not_2024"
        ]
        kpi2_row = rows_to_dicts(*run_query(conn, QUERY_KPI2))[0]
        kpi3_row = rows_to_dicts(*run_query(conn, QUERY_KPI3))[0]
        kpi4_row = rows_to_dicts(*run_query(conn, QUERY_KPI4_SUMMARY))[0]
        kpi4 = {
            "denom_orgs_2024_nonai_exclfmt": kpi4_row.get(
                "denom_orgs_2024_nonai_exclfmt"
            ),
            "num_orgs_new_ai_2025_exclfmt": kpi4_row.get(
                "num_orgs_new_ai_2025_exclfmt"
            ),
            "ratio_new_ai_2025_exclfmt": kpi4_row.get(
                "ratio_new_ai_2025_exclfmt"
            ),
        }
        kpi5_row = rows_to_dicts(
            *run_query(conn, QUERY_KPI5_COUNT_SHARE_2025_EXCLFMT)
        )[0]
        kpi5 = {
            "won_deal_count_total_2025_exclfmt": kpi5_row.get(
                "won_deal_count_total_2025_exclfmt"
            ),
            "won_deal_count_genai_2025_exclfmt": kpi5_row.get(
                "won_deal_count_genai_2025_exclfmt"
            ),
            "genai_deal_share_2025_exclfmt": kpi5_row.get(
                "genai_deal_share_2025_exclfmt"
            ),
        }
        kpi6_row = rows_to_dicts(
            *run_query(conn, QUERY_KPI6_DURATION_BUCKET_COUNTS_2025_GENAI_EXCLFMT)
        )[0]
        kpi6 = {
            "scope_deal_count": kpi6_row.get("scope_deal_count"),
            "valid_deal_count": kpi6_row.get("valid_deal_count"),
            "dropped_missing_or_invalid_dates": kpi6_row.get(
                "dropped_missing_or_invalid_dates"
            ),
            "cnt_1day": kpi6_row.get("cnt_1day"),
            "cnt_gt1_le1m": kpi6_row.get("cnt_gt1_le1m"),
            "cnt_ge1m_le3m": kpi6_row.get("cnt_ge1m_le3m"),
            "cnt_ge3m_le6m": kpi6_row.get("cnt_ge3m_le6m"),
            "cnt_ge6m_le12m": kpi6_row.get("cnt_ge6m_le12m"),
            "cnt_gt12m": kpi6_row.get("cnt_gt12m"),
        }

        # Summary structure
        summary = {
            "generated_at_utc": datetime.datetime.now(datetime.timezone.utc)
            .isoformat()
            .replace("+00:00", "Z"),
            "db_path": os.path.abspath(args.db_path),
            "kpi1": {
                "orgs_with_won_ever": kpi1_1,
                "orgs_won_2025_but_not_2024": kpi1_2,
            },
            "kpi2": kpi2_row,
            "kpi3": kpi3_row,
            "kpi4_2024_nonai_to_2025_new_ai_exclfmt": kpi4,
            "kpi5_2025_genai_deal_count_share_exclfmt": kpi5,
            "kpi6_2025_genai_duration_buckets_exclfmt": kpi6,
        }

        # stdout
        print("KPI1 orgs_with_won_ever:", kpi1_1)
        print("KPI1 orgs_won_2025_but_not_2024:", kpi1_2)
        print(
            "KPI2 genai_amount_sum:",
            kpi2_row.get("genai_amount_sum"),
            "total_amount_sum:",
            kpi2_row.get("total_amount_sum"),
            "genai_share:",
            kpi2_row.get("genai_share"),
            "deal_count_total:",
            kpi2_row.get("deal_count_total"),
            "deal_count_amount_missing:",
            kpi2_row.get("deal_count_amount_missing"),
        )
        print(
            "KPI3 denom_orgs_2024_nonai:",
            kpi3_row.get("denom_orgs_2024_nonai"),
            "num_orgs_new_ai_2025:",
            kpi3_row.get("num_orgs_new_ai_2025"),
            "ratio_new_ai_2025:",
            kpi3_row.get("ratio_new_ai_2025"),
        )
        print(
            "KPI4 denom_orgs_2024_nonai_exclfmt:",
            kpi4["denom_orgs_2024_nonai_exclfmt"],
            "num_orgs_new_ai_2025_exclfmt:",
            kpi4["num_orgs_new_ai_2025_exclfmt"],
            "ratio_new_ai_2025_exclfmt:",
            kpi4["ratio_new_ai_2025_exclfmt"],
        )
        share5 = kpi5["genai_deal_share_2025_exclfmt"]
        share5_pct = (share5 * 100.0) if share5 is not None else None
        print("\n[KPI5] 2025 Won(포맷 제외) 중 생성형AI 딜 '개수' 비중")
        print(
            f"  - denom(won_deal_count_total_2025_exclfmt): {kpi5['won_deal_count_total_2025_exclfmt']}"
        )
        print(
            f"  - num(won_deal_count_genai_2025_exclfmt):  {kpi5['won_deal_count_genai_2025_exclfmt']}"
        )
        if share5_pct is not None:
            print(
                f"  - share(genai_deal_share_2025_exclfmt):   {share5} ({share5_pct:.2f}%)"
            )
        else:
            print("  - share(genai_deal_share_2025_exclfmt):   NULL")
        print("\n[KPI6] 2025 GenAI Won(포맷 제외) 수강기간 버킷별 딜 개수")
        print(
            f"  scope_deal_count: {kpi6['scope_deal_count']}  | valid_deal_count: {kpi6['valid_deal_count']}  | dropped_missing_or_invalid_dates: {kpi6['dropped_missing_or_invalid_dates']}"
        )
        print(
            f"  buckets -> 1day:{kpi6['cnt_1day']}, 2-29:{kpi6['cnt_gt1_le1m']}, "
            f"30-89:{kpi6['cnt_ge1m_le3m']}, 90-179:{kpi6['cnt_ge3m_le6m']}, "
            f"180-365:{kpi6['cnt_ge6m_le12m']}, 366+:{kpi6['cnt_gt12m']}"
        )

        # JSON output
        with open(args.out_json, "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        # Audit datasets
        datasets = {
            "10_KPI1_WON_EVER_ORGS": QUERY_WON_EVER_ORGS,
            "11_KPI1_WON_EVER_DEALS": QUERY_WON_EVER_DEALS,
            "12_KPI1_NEW_2025_ORGS": QUERY_NEW_2025_ORGS,
            "13_KPI1_NEW_2025_DEALS": QUERY_NEW_2025_DEALS,
            "20_KPI2_SCOPE_DEALS_2025": QUERY_KPI2_SCOPE,
            "21_KPI2_GENAI_DEALS_2025": QUERY_KPI2_GENAI,
            "22_KPI2_DROPPED_AMOUNT_NULL": QUERY_KPI2_DROPPED_NULL,
            "30_KPI3_DENOM_ORGS_2024_NONAI": QUERY_KPI3_DENOM_ORGS,
            "31_KPI3_DENOM_DEALS_2024_WON": QUERY_KPI3_DENOM_DEALS,
            "32_KPI3_NUM_ORGS_NEW_AI_2025": QUERY_KPI3_NUM_ORGS,
            "33_KPI3_NUM_DEALS_AI_2025": QUERY_KPI3_NUM_DEALS,
            "40_KPI4_DENOM_ORGS_2024_NONAI_EXCLFMT": QUERY_KPI4_DENOM_ORGS,
            "41_KPI4_DENOM_DEALS_2024_WON_EXCLFMT": QUERY_KPI4_DENOM_DEALS_2024,
            "42_KPI4_NUM_ORGS_NEW_AI_2025_EXCLFMT": QUERY_KPI4_NUM_ORGS,
            "43_KPI4_NUM_DEALS_AI_2025_EXCLFMT": QUERY_KPI4_NUM_DEALS_2025_AI,
            "70_KPI6_VALID_DEALS_BUCKETED": QUERY_KPI6_DURATION_VALID_DEALS_2025_GENAI_EXCLFMT,
            "71_KPI6_DROPPED_DEALS_INVALID_DATES": QUERY_KPI6_DURATION_DROPPED_DEALS_2025_GENAI_EXCLFMT,
        }

        wb = Workbook()
        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        # README sheet
        ws_readme = wb.create_sheet("00_README")
        readme_rows = build_readme_sheet_data(
            args.db_path, summary["generated_at_utc"]
        )
        write_sheet(ws_readme, readme_rows, header_order=["item", "value"])

        # KPI summary sheet
        ws_summary = wb.create_sheet("01_KPI_SUMMARY")
        summary_rows = [
            {
                "kpi": "kpi1",
                "metric": "orgs_with_won_ever",
                "value": kpi1_1,
                "value_pct": None,
            },
            {
                "kpi": "kpi1",
                "metric": "orgs_won_2025_but_not_2024",
                "value": kpi1_2,
                "value_pct": None,
            },
            {
                "kpi": "kpi2",
                "metric": "genai_share",
                "value": kpi2_row.get("genai_share"),
                "value_pct": kpi2_row.get("genai_share"),
            },
            {
                "kpi": "kpi2",
                "metric": "genai_amount_sum",
                "value": kpi2_row.get("genai_amount_sum"),
                "value_pct": None,
            },
            {
                "kpi": "kpi2",
                "metric": "total_amount_sum",
                "value": kpi2_row.get("total_amount_sum"),
                "value_pct": None,
            },
            {
                "kpi": "kpi2",
                "metric": "deal_count_total",
                "value": kpi2_row.get("deal_count_total"),
                "value_pct": None,
            },
            {
                "kpi": "kpi2",
                "metric": "deal_count_amount_missing",
                "value": kpi2_row.get("deal_count_amount_missing"),
                "value_pct": None,
            },
            {
                "kpi": "kpi3",
                "metric": "ratio_new_ai_2025",
                "value": kpi3_row.get("ratio_new_ai_2025"),
                "value_pct": kpi3_row.get("ratio_new_ai_2025"),
            },
            {
                "kpi": "kpi3",
                "metric": "denom_orgs_2024_nonai",
                "value": kpi3_row.get("denom_orgs_2024_nonai"),
                "value_pct": None,
            },
            {
                "kpi": "kpi3",
                "metric": "num_orgs_new_ai_2025",
                "value": kpi3_row.get("num_orgs_new_ai_2025"),
                "value_pct": None,
            },
            {
                "kpi": "kpi4",
                "metric": "ratio_new_ai_2025_exclfmt",
                "value": kpi4["ratio_new_ai_2025_exclfmt"],
                "value_pct": kpi4["ratio_new_ai_2025_exclfmt"],
            },
            {
                "kpi": "kpi4",
                "metric": "denom_orgs_2024_nonai_exclfmt",
                "value": kpi4["denom_orgs_2024_nonai_exclfmt"],
                "value_pct": None,
            },
            {
                "kpi": "kpi4",
                "metric": "num_orgs_new_ai_2025_exclfmt",
                "value": kpi4["num_orgs_new_ai_2025_exclfmt"],
                "value_pct": None,
            },
            {
                "kpi": "kpi5",
                "metric": "genai_deal_share_2025_exclfmt",
                "value": kpi5["genai_deal_share_2025_exclfmt"],
                "value_pct": kpi5["genai_deal_share_2025_exclfmt"],
            },
            {
                "kpi": "kpi5",
                "metric": "won_deal_count_total_2025_exclfmt",
                "value": kpi5["won_deal_count_total_2025_exclfmt"],
                "value_pct": None,
            },
            {
                "kpi": "kpi5",
                "metric": "won_deal_count_genai_2025_exclfmt",
                "value": kpi5["won_deal_count_genai_2025_exclfmt"],
                "value_pct": None,
            },
            {
                "kpi": "kpi6",
                "metric": "scope_deal_count",
                "value": kpi6["scope_deal_count"],
                "value_pct": None,
            },
            {
                "kpi": "kpi6",
                "metric": "valid_deal_count",
                "value": kpi6["valid_deal_count"],
                "value_pct": None,
            },
            {
                "kpi": "kpi6",
                "metric": "dropped_missing_or_invalid_dates",
                "value": kpi6["dropped_missing_or_invalid_dates"],
                "value_pct": None,
            },
            {"kpi": "kpi6", "metric": "cnt_1day", "value": kpi6["cnt_1day"], "value_pct": None},
            {
                "kpi": "kpi6",
                "metric": "cnt_gt1_le1m",
                "value": kpi6["cnt_gt1_le1m"],
                "value_pct": None,
            },
            {
                "kpi": "kpi6",
                "metric": "cnt_ge1m_le3m",
                "value": kpi6["cnt_ge1m_le3m"],
                "value_pct": None,
            },
            {
                "kpi": "kpi6",
                "metric": "cnt_ge3m_le6m",
                "value": kpi6["cnt_ge3m_le6m"],
                "value_pct": None,
            },
            {
                "kpi": "kpi6",
                "metric": "cnt_ge6m_le12m",
                "value": kpi6["cnt_ge6m_le12m"],
                "value_pct": None,
            },
            {
                "kpi": "kpi6",
                "metric": "cnt_gt12m",
                "value": kpi6["cnt_gt12m"],
                "value_pct": None,
            },
        ]
        write_sheet(
            ws_summary,
            summary_rows,
            header_order=["kpi", "metric", "value", "value_pct"],
        )

        # Dataset sheets
        for sheet_name, sql in datasets.items():
            cols, rows = run_query(conn, sql)
            dict_rows = rows_to_dicts(cols, rows)
            dict_rows = append_org_names(dict_rows, org_map)
            ws = wb.create_sheet(sheet_name)
            header_order = list(cols)
            if org_map and "org_name" not in header_order and "org_id" in header_order:
                header_order.append("org_name")
            write_sheet(ws, dict_rows, header_order=header_order)

        wb.save(args.out_xlsx)
    finally:
        conn.close()


if __name__ == "__main__":
    try:
        main()
    except sqlite3.Error as exc:
        sys.stderr.write(f"SQLite error: {exc}\n")
        sys.exit(1)
