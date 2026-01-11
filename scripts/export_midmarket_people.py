import argparse
import re
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# README-style header:
# - Filters organizations by inferred sizeGroup (default: 중견기업) using dashboard.server.database.infer_size_group when available.
# - Aggregates Won deal amounts by year (contract_date → fallback any 20xx in other date fields).
# - Merges people info from lead (priority), memo (raw text key:value parse), people table (fallback).
# - Detects consent/phone/email/job/title/sequence from lead columns (auto keyword match) or memo text keys.
# - Supports stub people from deals if missing in people table.
# - Writes Excel with midmarket_people sheet; won amounts kept in 원 단위 with numeric formatting.
# Usage example:
#   python scripts/export_midmarket_people.py --db-path salesmap_latest.db --size-group 중견기업 --years 2023,2024,2025 --out exports/midmarket_people.xlsx

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.append(str(REPO_ROOT))

try:
    from dashboard.server.database import infer_size_group as _infer_size_group
except Exception:
    _infer_size_group = None


def get_infer_size_group_fallback():
    if _infer_size_group:
        return _infer_size_group

    def _fallback(name: str, size_raw: str) -> str:
        text = f"{name or ''} {size_raw or ''}"
        if "중견" in text:
            return "중견기업"
        if "대기업" in text:
            return "대기업"
        if "중소" in text:
            return "중소기업"
        if "공공" in text:
            return "공공기관"
        if "대학" in text:
            return "대학교"
        return "기타/미입력"

    return _fallback


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export midmarket people info to Excel.")
    parser.add_argument("--db-path", default="salesmap_latest.db", help="Path to SQLite snapshot")
    parser.add_argument("--size-group", default="중견기업", help="Target size group (default: 중견기업)")
    parser.add_argument(
        "--years", default="2023,2024,2025", help="Comma-separated years for Won sums (default: 2023,2024,2025)"
    )
    parser.add_argument(
        "--out", default="export_midmarket_people_2023_2025.xlsx", help="Output Excel path (default: export_midmarket_people_2023_2025.xlsx)"
    )
    return parser.parse_args()


def get_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    rows = conn.execute(f"PRAGMA table_info('{table}')").fetchall()
    return [row[1] for row in rows]


def ensure_tables(conn: sqlite3.Connection, tables: Sequence[str]) -> None:
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    )
    existing = {row[0] for row in cur.fetchall()}
    missing = [t for t in tables if t not in existing]
    if missing:
        raise SystemExit(f"Missing required table(s): {', '.join(missing)}")


def parse_year_from_fields(contract_date: Any, alt_dates: Sequence[Any]) -> Optional[str]:
    def pick_year(val: Any) -> Optional[str]:
        if val is None:
            return None
        text = str(val)
        if len(text) >= 4 and text[:4].isdigit():
            return text[:4]
        m = re.search(r"(20\d{2})", text)
        if m:
            return m.group(1)
        return None

    year = pick_year(contract_date)
    if year:
        return year
    for val in alt_dates:
        year = pick_year(val)
        if year:
            return year
    return None


def to_number(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def parse_amount(amount_raw: Any, expected_raw: Any) -> float:
    amt = to_number(amount_raw)
    if amt is not None and amt > 0:
        return amt
    exp = to_number(expected_raw)
    if exp is not None and exp > 0:
        return exp
    return 0.0


def pick_first(cols: Sequence[str], keywords: Sequence[str]) -> Optional[str]:
    lowered = [(c, c.lower()) for c in cols]
    for kw in keywords:
        for orig, low in lowered:
            if kw in low:
                return orig
    return None


def safe_str(val: Any) -> str:
    return str(val).strip() if val is not None else ""


def load_orgs(conn: sqlite3.Connection, target_size: str, infer_size_group) -> Dict[str, Dict[str, Any]]:
    cols = get_columns(conn, "organization")
    needed = {
        "id": "id",
        "name": "이름" if "이름" in cols else "id",
        "size": "기업 규모" if "기업 규모" in cols else None,
        "industry_major": "업종 구분(대)" if "업종 구분(대)" in cols else None,
        "industry_mid": "업종 구분(중)" if "업종 구분(중)" in cols else None,
        "phone": "전화" if "전화" in cols else None,
    }
    def q(col: str) -> str:
        return f'"{col}"'

    query_cols = [f'{q(needed["id"])} AS id', f'COALESCE({q(needed["name"])}, {q(needed["id"])}) AS name']
    if needed["size"]:
        query_cols.append(f'{q(needed["size"])} AS size')
    if needed["industry_major"]:
        query_cols.append(f'{q(needed["industry_major"])} AS industry_major')
    if needed["industry_mid"]:
        query_cols.append(f'{q(needed["industry_mid"])} AS industry_mid')
    if needed["phone"]:
        query_cols.append(f'{q(needed["phone"])} AS phone')
    query = f"SELECT {', '.join(query_cols)} FROM organization"
    rows = conn.execute(query).fetchall()
    orgs: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        org_id = safe_str(row["id"])
        size_raw = row["size"] if "size" in row.keys() else None
        size_group = infer_size_group(row["name"], size_raw)
        if size_group != target_size:
            continue
        orgs[org_id] = {
            "orgId": org_id,
            "orgName": row["name"],
            "sizeRaw": size_raw or "",
            "sizeGroup": size_group,
            "industryMajor": row["industry_major"] if "industry_major" in row.keys() else "",
            "industryMid": row["industry_mid"] if "industry_mid" in row.keys() else "",
            "orgPhone": row["phone"] if "phone" in row.keys() else "",
        }
    return orgs


def load_won_totals_and_people(conn: sqlite3.Connection, org_ids: Sequence[str], years: Sequence[str]) -> Tuple[Dict[str, Dict[str, float]], Dict[str, set]]:
    if not org_ids:
        return {}, {}
    cols = get_columns(conn, "deal")
    status_col = "상태" if "상태" in cols else None
    amount_col = "금액" if "금액" in cols else None
    expected_col = "예상 체결액" if "예상 체결액" in cols else None
    contract_col = "계약 체결일" if "계약 체결일" in cols else None
    expected_date_col = "수주 예정일" if "수주 예정일" in cols else None
    if not status_col or not amount_col:
        raise SystemExit("deal table must have 상태 and 금액 columns")

    placeholders = ",".join("?" for _ in org_ids)
    def q(col: str) -> str:
        return f'"{col}"'

    query_cols = [
        "organizationId AS orgId",
        f"{q(status_col)} AS status",
        f"{q(amount_col)} AS amount",
    ]
    if expected_col:
        query_cols.append(f"{q(expected_col)} AS expected_amount")
    else:
        query_cols.append("NULL AS expected_amount")
    query_cols.append(f"{q(contract_col)} AS contract_date" if contract_col else "NULL AS contract_date")
    query_cols.append(f"{q(expected_date_col)} AS expected_date" if expected_date_col else "NULL AS expected_date")
    query_cols.append("peopleId AS peopleId")
    query = f"SELECT {', '.join(query_cols)} FROM deal WHERE organizationId IN ({placeholders})"
    rows = conn.execute(query, tuple(org_ids)).fetchall()

    totals: Dict[str, Dict[str, float]] = {org: {y: 0.0 for y in years} for org in org_ids}
    people_map: Dict[str, set] = defaultdict(set)
    for row in rows:
        org_id = safe_str(row["orgId"])
        status = safe_str(row["status"])
        if status != "Won":
            continue
        year = parse_year_from_fields(row["contract_date"], [row["expected_date"]])
        if not year or year not in years:
            continue
        amt = parse_amount(row["amount"], row["expected_amount"])
        totals.setdefault(org_id, {y: 0.0 for y in years})
        totals[org_id][year] = totals[org_id].get(year, 0.0) + amt
        if row["peopleId"]:
            people_map[org_id].add(safe_str(row["peopleId"]))
    return totals, people_map


def load_people(conn: sqlite3.Connection, org_ids: Sequence[str]) -> Dict[str, Dict[str, Any]]:
    cols = get_columns(conn, "people")
    id_col = "id"
    org_col = "organizationId"
    name_col = "이름" if "이름" in cols else "id"
    email_col = pick_first(cols, ["email", "이메일"])
    phone_col = pick_first(cols, ["phone", "전화"])
    job_col = pick_first(cols, ["담당업무", "담당 업무"])
    title_col = pick_first(cols, ["직급", "직책"])
    def q(col: str) -> str:
        return f'"{col}"'

    query_cols = [
        f"{q(id_col)} AS id",
        f"{q(org_col)} AS organizationId",
        f"COALESCE({q(name_col)}, {q(id_col)}) AS name",
    ]
    if title_col:
        query_cols.append(f"{q(title_col)} AS title")
    if job_col:
        query_cols.append(f"{q(job_col)} AS job")
    if email_col:
        query_cols.append(f"{q(email_col)} AS email")
    if phone_col:
        query_cols.append(f"{q(phone_col)} AS phone")
    query = f"SELECT {', '.join(query_cols)} FROM people WHERE {org_col} IN ({','.join('?' for _ in org_ids)})"
    rows = conn.execute(query, tuple(org_ids)).fetchall()
    people: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        pid = safe_str(row["id"])
        org_id = safe_str(row["organizationId"])
        people[pid] = {
            "personId": pid,
            "orgId": org_id,
            "personName": row["name"],
            "personTitle": safe_str(row["title"]) if "title" in row.keys() else "",
            "personJob": safe_str(row["job"]) if "job" in row.keys() else "",
            "personEmail": safe_str(row["email"]) if "email" in row.keys() else "",
            "personPhone": safe_str(row["phone"]) if "phone" in row.keys() else "",
        }
    return people


def load_leads(conn: sqlite3.Connection, org_ids: Sequence[str]) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    existing = get_columns(conn, "lead")
    if not existing:
        return {}, {}
    pid_col = "peopleId" if "peopleId" in existing else None
    org_col = "organizationId" if "organizationId" in existing else None
    if not pid_col and not org_col:
        return {}, {}

    email_col = pick_first(existing, ["email", "이메일"])
    phone_col = pick_first(existing, ["phone", "전화"])
    consent_col = pick_first(existing, ["동의", "consent"])
    seq_col = pick_first(existing, ["sequence", "시퀀스"])
    title_col = pick_first(existing, ["직급", "직책"])
    job_col = pick_first(existing, ["담당업무", "담당 업무"])
    ts_cols = [c for c in existing if "updated" in c.lower()] or [c for c in existing if "created" in c.lower()]

    def q(col: str) -> str:
        return f'"{col}"'

    def qa(col: str) -> str:
        return f'{q(col)} AS "{col}"'

    query_cols = ["id"]
    for col in [pid_col, org_col, email_col, phone_col, consent_col, seq_col, title_col, job_col]:
        if col:
            query_cols.append(qa(col))
    for col in ts_cols:
        query_cols.append(qa(col))
    query = f"SELECT {', '.join(query_cols)} FROM lead"
    rows = conn.execute(query).fetchall()

    by_pid: Dict[str, Dict[str, Any]] = {}
    by_org: Dict[str, Dict[str, Any]] = {}

    def pick_ts(row) -> Optional[str]:
        best = None
        for col in ts_cols:
            if col in row.keys() and row[col]:
                best = row[col]
                break
        return best

    for row in rows:
        pid = safe_str(row[pid_col]) if pid_col and pid_col in row.keys() and row[pid_col] is not None else ""
        oid = safe_str(row[org_col]) if org_col and org_col in row.keys() and row[org_col] is not None else ""
        if oid and oid not in org_ids:
            continue
        payload = {
            "email": safe_str(row[email_col]) if email_col and email_col in row.keys() else "",
            "phone": safe_str(row[phone_col]) if phone_col and phone_col in row.keys() else "",
            "marketingConsent": safe_str(row[consent_col]) if consent_col and consent_col in row.keys() else "",
            "inSequence": safe_str(row[seq_col]) if seq_col and seq_col in row.keys() else "",
            "personTitle": safe_str(row[title_col]) if title_col and title_col in row.keys() else "",
            "personJob": safe_str(row[job_col]) if job_col and job_col in row.keys() else "",
            "lastContactSource": "lead",
            "lastContactAt": safe_str(pick_ts(row) or ""),
        }
        if pid:
            existing_payload = by_pid.get(pid)
            if not existing_payload or (payload["lastContactAt"] and payload["lastContactAt"] > existing_payload.get("lastContactAt", "")):
                by_pid[pid] = payload
        elif oid:
            existing_payload = by_org.get(oid)
            if not existing_payload or (payload["lastContactAt"] and payload["lastContactAt"] > existing_payload.get("lastContactAt", "")):
                by_org[oid] = payload
    return by_pid, by_org


def parse_memo_text(text: str) -> Dict[str, str]:
    fields: Dict[str, str] = {}
    for raw in text.splitlines():
        m = re.match(r"^\s*([^\:]+):\s*(.+)$", raw)
        if not m:
            continue
        key = m.group(1).strip().replace(" ", "").lower()
        val = m.group(2).strip()
        if not val:
            continue
        fields[key] = val
    return fields


def extract_from_memo(raw_text: str) -> Dict[str, str]:
    parsed = parse_memo_text(raw_text)
    out: Dict[str, str] = {}
    for k, v in parsed.items():
        if "마케팅수신동의" in k or "고객마케팅수신동의" in k or "동의" in k:
            out["marketingConsent"] = v
        elif "전화" in k:
            out["personPhone"] = v
        elif "이메일" in k:
            out["personEmail"] = v
        elif "담당업무" in k:
            out["personJob"] = v
        elif "직급" in k or "직책" in k:
            out["personTitle"] = v
        elif "시퀀스" in k:
            out["inSequence"] = v
    return out


def load_memos(conn: sqlite3.Connection, org_ids: Sequence[str]) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    cols = get_columns(conn, "memo")
    if not cols:
        return {}, {}
    text_col = "text" if "text" in cols else None
    created_col = "createdAt" if "createdAt" in cols else None
    if not text_col:
        return {}, {}
    query = (
        f"SELECT id, {text_col} AS text, "
        f"{'peopleId' if 'peopleId' in cols else 'NULL'} AS peopleId, "
        f"{'organizationId' if 'organizationId' in cols else 'NULL'} AS organizationId, "
        f"{created_col if created_col else 'NULL'} AS createdAt "
        f"FROM memo"
    )
    rows = conn.execute(query).fetchall()
    by_pid: Dict[str, Dict[str, Any]] = {}
    by_org: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        pid = safe_str(row["peopleId"]) if "peopleId" in row.keys() else ""
        oid = safe_str(row["organizationId"]) if "organizationId" in row.keys() else ""
        if oid and oid not in org_ids:
            continue
        text = row["text"] or ""
        parsed = extract_from_memo(text)
        if not parsed:
            continue
        payload = {
            **parsed,
            "lastContactSource": "memo",
            "lastContactAt": safe_str(row["createdAt"]) if "createdAt" in row.keys() else "",
        }
        if pid:
            existing = by_pid.get(pid)
            if not existing or payload["lastContactAt"] > existing.get("lastContactAt", ""):
                by_pid[pid] = payload
        elif oid:
            existing = by_org.get(oid)
            if not existing or payload["lastContactAt"] > existing.get("lastContactAt", ""):
                by_org[oid] = payload
    return by_pid, by_org


def merge_contact(base: Dict[str, Any], lead: Optional[Dict[str, Any]], memo: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    result = dict(base)

    def pick(field: str) -> str:
        for src in (lead, memo):
            if src and src.get(field):
                return src[field]
        return result.get(field, "") or ""

    for field in ["personEmail", "personPhone", "personJob", "personTitle", "marketingConsent", "inSequence"]:
        result[field] = pick(field)

    # last contact
    if lead and lead.get("lastContactAt"):
        result["lastContactSource"] = "lead"
        result["lastContactAt"] = lead.get("lastContactAt", "")
    elif memo and memo.get("lastContactAt"):
        result["lastContactSource"] = "memo"
        result["lastContactAt"] = memo.get("lastContactAt", "")
    else:
        result["lastContactSource"] = ""
        result["lastContactAt"] = ""
    return result


def build_rows(orgs: Dict[str, Dict[str, Any]], people: Dict[str, Dict[str, Any]], won_people: Dict[str, set], totals: Dict[str, Dict[str, float]], lead_by_pid: Dict[str, Dict[str, Any]], lead_by_org: Dict[str, Dict[str, Any]], memo_by_pid: Dict[str, Dict[str, Any]], memo_by_org: Dict[str, Dict[str, Any]], years: Sequence[str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for org_id, org_meta in orgs.items():
        people_ids = set(pid for pid, pdata in people.items() if pdata.get("orgId") == org_id)
        people_ids.update(won_people.get(org_id, set()))
        if not people_ids:
            continue
        won = totals.get(org_id, {y: 0.0 for y in years})
        for pid in sorted(people_ids):
            base = people.get(pid, {"personId": pid, "orgId": org_id, "personName": pid, "personTitle": "", "personJob": "", "personEmail": "", "personPhone": ""})
            base.setdefault("personName", pid)
            lead = lead_by_pid.get(pid) or lead_by_org.get(org_id)
            memo = memo_by_pid.get(pid) or memo_by_org.get(org_id)
            merged = merge_contact(base, lead, memo)
            row = {
                **org_meta,
                "wonAmount2023": int(won.get("2023", 0.0)),
                "wonAmount2024": int(won.get("2024", 0.0)),
                "wonAmount2025": int(won.get("2025", 0.0)),
                "wonAmountTotal2325": int(sum(won.values())),
                **merged,
            }
            rows.append(row)
    return rows


def write_excel(rows: List[Dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "midmarket_people"
    headers = [
        "orgId",
        "orgName",
        "sizeRaw",
        "sizeGroup",
        "industryMajor",
        "industryMid",
        "orgPhone",
        "wonAmount2023",
        "wonAmount2024",
        "wonAmount2025",
        "wonAmountTotal2325",
        "personId",
        "personName",
        "personTitle",
        "personJob",
        "personEmail",
        "personPhone",
        "inSequence",
        "marketingConsent",
        "lastContactSource",
        "lastContactAt",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    won_cols = [headers.index("wonAmount2023") + 1, headers.index("wonAmount2024") + 1, headers.index("wonAmount2025") + 1, headers.index("wonAmountTotal2325") + 1]
    for col_idx in won_cols:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                c.number_format = "#,##0"

    for i, header in enumerate(headers, start=1):
        width = max(len(header), 12)
        ws.column_dimensions[get_column_letter(i)].width = width

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[export] Wrote {out_path}")


def main():
    args = parse_args()
    years = [y.strip() for y in args.years.split(",") if y.strip()]
    infer_fn = get_infer_size_group_fallback()
    db_path = Path(args.db_path)
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        ensure_tables(conn, ["organization", "people", "deal", "memo"])
        orgs = load_orgs(conn, args.size_group, infer_fn)
        if not orgs:
            print(f"[export] No organizations found for sizeGroup={args.size_group}")
            return
        totals, won_people = load_won_totals_and_people(conn, list(orgs.keys()), years)
        people = load_people(conn, list(orgs.keys()))
        lead_by_pid, lead_by_org = load_leads(conn, list(orgs.keys()))
        memo_by_pid, memo_by_org = load_memos(conn, list(orgs.keys()))
        rows = build_rows(orgs, people, won_people, totals, lead_by_pid, lead_by_org, memo_by_pid, memo_by_org, years)
        out_path = Path(args.out)
        write_excel(rows, out_path)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
