import argparse
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

import pandas as pd

# Ensure repository root is importable
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.append(str(REPO_ROOT))

# Reuse existing DB logic (grade calculation, constants, helpers)
from dashboard.server import database as db

TIERS = ["S0", "P0", "P1", "P2"]
HEADERS = [
    "티어(S0, P0, P1, P2)",
    "organization id",
    "organization 이름",
    "people의 id",
    "people의 이름",
    "people의 소속 상위 조직",
    "people의 팀(명함/메일서명)",
    "people의 직급(명함/메일서명)",
    "people의 '담당 교육 영역'",
    "people의 '담당자'의 id",
    "people의 '담당자'의 name",
    "2024 딜 개수",
    "2025 딜 개수",
    "2024 딜 카드 예시",
    "2025 딜카드 예시",
]

BASE_URL = (os.getenv("SALESMAP_WEB_BASE") or "https://salesmap.kr/64cb5beda5a78ae225d7815b").rstrip(
    "/"
)
PEOPLE_URL_TMPL = os.getenv(
    "SALESMAP_PEOPLE_URL_TMPL", "{base}/contact/people/{people_id}"
)
DEAL_URL_TMPL = os.getenv("SALESMAP_DEAL_URL_TMPL", "{base}/deal/{deal_id}")
DATE_FLOOR = "2024-01-01"


def _log_people_schema(conn) -> None:
    rows = conn.execute("PRAGMA table_info('people')").fetchall()
    cols = [row[1] for row in rows]
    print(f"[export] people columns: {cols}")


def _log_deal_schema(conn) -> List[str]:
    rows = conn.execute("PRAGMA table_info('deal')").fetchall()
    cols = [row[1] for row in rows]
    print(f"[export] deal columns: {cols}")
    return cols


def _clean_text(val: Any, allow_empty: bool = False, blank_if_missing: bool = False) -> str:
    if val is None:
        if blank_if_missing or allow_empty:
            return ""
        return "미입력"
    text = str(val).strip()
    if not text:
        if blank_if_missing or allow_empty:
            return ""
        return "미입력"
    return text


def _parse_owner(raw: Any) -> Tuple[str, str]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return "", "미입력"
    loaded = db._safe_json_load(raw)  # type: ignore[attr-defined]
    candidate = loaded[0] if isinstance(loaded, list) and loaded else loaded

    def pick(obj: Any) -> Tuple[str, str]:
        if isinstance(obj, dict):
            oid = _clean_text(obj.get("id"), allow_empty=True)
            name = _clean_text(obj.get("name"), allow_empty=False)
            return oid, name
        if obj is None:
            return "", "미입력"
        text = str(obj).strip()
        if not text:
            return "", "미입력"
        return "", text

    oid, name = pick(candidate)
    if name == "미입력" and isinstance(raw, str):
        name = raw.strip() or "미입력"
    return oid, name


def _fetch_people_for_orgs(conn, org_ids: Sequence[str]) -> Dict[str, List[Dict[str, Any]]]:
    placeholders = ",".join("?" for _ in org_ids)
    query = (
        'SELECT id, organizationId, COALESCE("이름", id) AS name, '
        '"소속 상위 조직" AS upper_org, "팀(명함/메일서명)" AS team_signature, '
        '"직급(명함/메일서명)" AS title_signature, "담당 교육 영역" AS edu_area, '
        '"담당자" AS owner_json '
        "FROM people "
        f"WHERE organizationId IN ({placeholders})"
    )
    rows = conn.execute(query, org_ids).fetchall()
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(row["organizationId"], []).append(dict(row))
    return grouped


def _row_sort_key(row: Dict[str, Any]):
    tier_order = {t: i for i, t in enumerate(TIERS)}
    upper = row["people의 소속 상위 조직"]
    upper_missing_flag = 1 if not upper or upper == "미입력" else 0
    upper_for_sort = upper or "미입력"
    return (
        tier_order.get(row["티어(S0, P0, P1, P2)"], len(TIERS)),
        row["organization 이름"],
        upper_missing_flag,
        upper_for_sort,
        row["people의 '담당자'의 name"],
        row["people의 이름"],
        row["people의 id"],
    )


def _sort_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(rows, key=_row_sort_key)


def _resolve_column(columns: List[str], candidates: Sequence[str]) -> str:
    for cand in candidates:
        if cand in columns:
            return cand
    raise KeyError(f"None of candidates {candidates} found in columns {columns}")


def _q(col: str) -> str:
    return f'"{col}"'


def _normalize_date10(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    if len(s) >= 10:
        return s[:10]
    return s


def _extract_year(val: str) -> str:
    return val[:4] if len(val) >= 4 else ""


def _build_people_link(people_id: str) -> str | None:
    pid = (people_id or "").strip()
    if not pid:
        return None
    return PEOPLE_URL_TMPL.format(base=BASE_URL, people_id=pid)


def _build_deal_link(deal_id: str) -> str | None:
    did = (deal_id or "").strip()
    if not did:
        return None
    return DEAL_URL_TMPL.format(base=BASE_URL, deal_id=did)


def export_rank2025_tier_people(db_path: Path, out_path: Path, size: str) -> None:
    if not db_path.exists():
        raise FileNotFoundError(f"Database not found at {db_path}")

    os.makedirs(out_path.parent, exist_ok=True)
    items = db.get_rank_2025_deals(size=size, db_path=db_path)
    filtered = [item for item in items if item.get("grade") in TIERS]
    org_ids = [item["orgId"] for item in filtered]
    if not org_ids:
        print("[export] No organizations matched the requested tiers.")
        return

    with db._connect(db_path) as conn:  # type: ignore[attr-defined]
        _log_people_schema(conn)
        deal_cols = _log_deal_schema(conn)
        name_col = _resolve_column(deal_cols, ["이름", "name"])
        created_col = _resolve_column(deal_cols, ["생성 날짜", "createdAt", "created_at"])
        deal_rows = conn.execute(
            f"SELECT id AS deal_id, peopleId AS people_id, "
            f"COALESCE({_q(name_col)}, id) AS deal_name, "
            f"{_q(created_col)} AS created_raw "
            "FROM deal "
            "WHERE peopleId IS NOT NULL"
        ).fetchall()

        eligible_people: set[str] = set()
        stats_by_people: Dict[str, Dict[str, Any]] = {}
        for drow in deal_rows:
            pid = (drow["people_id"] or "").strip()
            if not pid:
                continue
            date10 = _normalize_date10(drow["created_raw"])
            if not date10:
                continue
            if date10 < DATE_FLOOR:
                continue
            year = _extract_year(date10)
            deal_name = _clean_text(drow["deal_name"])
            deal_id = _clean_text(drow["deal_id"], allow_empty=True)
            stats = stats_by_people.setdefault(
                pid,
                {
                    "cnt_2024": 0,
                    "cnt_2025": 0,
                    "cand_2024": [],
                    "cand_2025": [],
                },
            )
            eligible_people.add(pid)
            if year == "2024":
                stats["cnt_2024"] += 1
                stats["cand_2024"].append((deal_name, deal_id))
            elif year == "2025":
                stats["cnt_2025"] += 1
                stats["cand_2025"].append((deal_name, deal_id))

        people_map = _fetch_people_for_orgs(conn, org_ids)

    def pick_example(cands: List[Tuple[str, str]]) -> Tuple[str, str]:
        if not cands:
            return "딜 부재", ""
        cands_sorted = sorted(cands, key=lambda x: (x[0], x[1]), reverse=True)
        return cands_sorted[0]

    entries: List[Tuple[Dict[str, Any], Dict[str, Any]]] = []
    for org in filtered:
        org_id = org["orgId"]
        org_name = _clean_text(org.get("orgName"))
        tier = org.get("grade") or ""
        people_list = people_map.get(org_id, [])
        for person in people_list:
            pid = _clean_text(person.get("id"), allow_empty=True)
            if not pid or pid not in eligible_people:
                continue
            stats = stats_by_people.get(pid, {})
            ex24_name, ex24_id = pick_example(stats.get("cand_2024", []))
            ex25_name, ex25_id = pick_example(stats.get("cand_2025", []))
            owner_id, owner_name = _parse_owner(person.get("owner_json"))
            row = {
                "티어(S0, P0, P1, P2)": tier,
                "organization id": _clean_text(org_id, allow_empty=True),
                "organization 이름": org_name,
                "people의 id": pid,
                "people의 이름": _clean_text(person.get("name")),
                "people의 소속 상위 조직": _clean_text(
                    person.get("upper_org"), blank_if_missing=True
                ),
                "people의 팀(명함/메일서명)": _clean_text(
                    person.get("team_signature"), blank_if_missing=True
                ),
                "people의 직급(명함/메일서명)": _clean_text(
                    person.get("title_signature"), blank_if_missing=True
                ),
                "people의 '담당 교육 영역'": _clean_text(
                    person.get("edu_area"), blank_if_missing=True
                ),
                "people의 '담당자'의 id": owner_id,
                "people의 '담당자'의 name": owner_name,
                "2024 딜 개수": stats.get("cnt_2024", 0),
                "2025 딜 개수": stats.get("cnt_2025", 0),
                "2024 딜 카드 예시": ex24_name,
                "2025 딜카드 예시": ex25_name,
            }
            link_info = {
                "people_id": pid,
                "deal24_id": ex24_id,
                "deal25_id": ex25_id,
            }
            entries.append((row, link_info))

    entries_sorted = sorted(entries, key=lambda pair: _row_sort_key(pair[0]))
    rows = [r for r, _ in entries_sorted]
    link_meta = [meta for _, meta in entries_sorted]

    df = pd.DataFrame(rows, columns=HEADERS)
    df.to_excel(out_path, index=False, engine="openpyxl")

    # Post-processing: freeze header, autofilter, column width, hyperlinks, highlighting
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(out_path)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_map = {cell.value: cell.column_letter for cell in ws[1]}
    people_col = header_map.get("people의 이름")
    deal24_col = header_map.get("2024 딜 카드 예시")
    deal25_col = header_map.get("2025 딜카드 예시")
    highlight_cols = [
        header_map.get("people의 소속 상위 조직"),
        header_map.get("people의 팀(명함/메일서명)"),
        header_map.get("people의 직급(명함/메일서명)"),
        header_map.get("people의 '담당 교육 영역'"),
    ]
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for idx, meta in enumerate(link_meta, start=2):  # Excel rows start at 2 (after header)
        if people_col:
            cell = ws[f"{people_col}{idx}"]
            url = _build_people_link(meta.get("people_id", ""))
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"
        if deal24_col and meta.get("deal24_id"):
            cell = ws[f"{deal24_col}{idx}"]
            url = _build_deal_link(meta["deal24_id"])
            if url and cell.value != "딜 부재":
                cell.hyperlink = url
                cell.style = "Hyperlink"
        if deal25_col and meta.get("deal25_id"):
            cell = ws[f"{deal25_col}{idx}"]
            url = _build_deal_link(meta["deal25_id"])
            if url and cell.value != "딜 부재":
                cell.hyperlink = url
                cell.style = "Hyperlink"
        for col_letter in highlight_cols:
            if not col_letter:
                continue
            c = ws[f"{col_letter}{idx}"]
            if c.value is None or str(c.value).strip() == "":
                c.fill = yellow

    for col_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)
    wb.save(out_path)
    print(
        f"[export] tier orgs: {len(filtered)}, people (eligible rows): {len(rows)}, written: {out_path}"
    )


def main():
    parser = argparse.ArgumentParser(description="Export S0/P0/P1/P2 tier org people to XLSX.")
    parser.add_argument("--db-path", default="./salesmap_latest.db", help="Path to SQLite DB")
    parser.add_argument(
        "--out", default="./exports/rank2025_tier_people.xlsx", help="Output XLSX path"
    )
    parser.add_argument(
        "--size",
        default="전체",
        help='Organization size filter (e.g., "대기업", "전체") applied to rank-2025',
    )
    args = parser.parse_args()
    export_rank2025_tier_people(Path(args.db_path), Path(args.out), size=args.size)


if __name__ == "__main__":
    main()
