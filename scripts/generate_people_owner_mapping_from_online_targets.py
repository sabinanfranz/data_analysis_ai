#!/usr/bin/env python3
"""
Generate people-owner mapping workbooks from counterparty targets sheets and salesmap DB.

Outputs (written to dashboard/server/resources/outputs/):
  - people_owner_mapping_from_26_online_targets.xlsx
  - people_owner_mapping_from_26_lecture_targets.xlsx

Each workbook contains:
  - result: mapped people rows
  - log: skipped rows and matching notes

Usage: python scripts/generate_people_owner_mapping_from_online_targets.py
"""
from __future__ import annotations

import json
import sqlite3
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


# PART_STRUCTURE extracted from org_tables_v2.html (source of truth).
PART_STRUCTURE: Dict[str, Dict[str, List[str]]] = {
    "기업교육 1팀": {
        "1파트": ["김솔이", "황초롱", "김정은", "김동찬", "정태윤", "서정연", "오진선", "공새봄", "김별"],
        "2파트": ["강지선", "정하영", "박범규", "하승민", "이은서", "김세연"],
    },
    "기업교육 2팀": {
        "1파트": ["권노을", "이윤지B", "이현진", "김민선", "강연정", "방신우", "홍제환", "정선희"],
        "2파트": ["정다혜", "임재우", "송승희", "손승완", "김윤지", "손지훈", "홍예진"],
        "온라인셀": ["강진우", "강다현", "이수빈"],
    },
    "공공교육팀": {
        "전체": ["이준석", "김미송", "오정민", "조경원", "김다인", "서민정", "김지원", "김진호"],
    },
}


ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "dashboard/server/resources/counterparty_targets_2026.xlsx"
DB_PATH = ROOT / "dashboard/server/resources/salesmap_latest.db"
if not DB_PATH.exists():
    DB_PATH = ROOT / "salesmap_latest.db"

OUTPUT_DIR = ROOT / "dashboard/server/resources/outputs"
RESULT_SHEET = "result"
LOG_SHEET = "log"

SHEETS = [
    ("26 온라인 타겟", "people_owner_mapping_from_26_online_targets.xlsx"),
    ("26 출강 타겟", "people_owner_mapping_from_26_lecture_targets.xlsx"),
]

# caches populated in main()
SOURCE_WB = None
ORG_INDEX: Optional[Dict[str, List[str]]] = None
PEOPLE_BY_ORG: Optional[Dict[str, List[sqlite3.Row]]] = None
NAME_TO_TEAM: Optional[Dict[str, List[str]]] = None


def clean_str(value: Optional[object]) -> str:
    return "" if value is None else str(value).strip()


def norm_upper_org(value: Optional[object]) -> str:
    text = clean_str(value)
    return text if text else "미입력"


def parse_assignee(cell_value: Optional[object]) -> Tuple[Optional[str], Optional[str], int]:
    """
    Returns (assignee, error_reason, token_count).

    error_reason is one of:
      - assignee_empty: empty/blank
      - assignee_multi: more than one token
    """
    text = clean_str(cell_value)
    if not text:
        return "", "assignee_empty", 0

    tokens = [token.strip() for token in text.split(",") if token.strip()]
    if len(tokens) == 0:
        return "", "assignee_empty", 0
    if len(tokens) > 1:
        return None, "assignee_multi", len(tokens)
    return tokens[0], None, 1


def normalize_owner_name(raw_owner: Optional[object]) -> Tuple[str, bool, bool, str]:
    """
    Normalize owner string to just the human name.

    Returns (normalized_name, was_json_candidate, json_name_extracted, parse_error_msg).
    """
    text = clean_str(raw_owner)
    if not text:
        return "", False, False, ""

    stripped = text.strip()
    is_json_candidate = stripped.startswith("{") and stripped.endswith("}")
    json_name_extracted = False
    parse_error = ""

    if is_json_candidate:
        try:
            parsed = json.loads(stripped)
            if isinstance(parsed, dict):
                candidate = parsed.get("name") or parsed.get("displayName")
                if candidate and clean_str(candidate):
                    return clean_str(candidate), True, True, ""
            # fall through to general handling when name missing
        except Exception as exc:  # keep broad to avoid skipping due to ValueError TypeError etc.
            parse_error = f"{exc.__class__.__name__}: {exc}"
        else:
            # JSON parsed but name missing -> fall back
            json_name_extracted = False

    # non-JSON or fallback path
    text_no_extra = stripped
    if "외" in text_no_extra:
        text_no_extra = text_no_extra.split("외", 1)[0].strip()
    for sep in ("(", "（"):
        if sep in text_no_extra:
            text_no_extra = text_no_extra.split(sep, 1)[0].strip()
            break
    return text_no_extra, is_json_candidate, json_name_extracted, parse_error


# minimal assertions to guard key behaviors
assert normalize_owner_name('{"id":"x","name":"정다혜"}')[0] == "정다혜"
assert normalize_owner_name("정다혜")[0] == "정다혜"
assert normalize_owner_name("정다혜 외 1명")[0] == "정다혜"
assert normalize_owner_name("")[0] == ""


def build_org_index(conn: sqlite3.Connection) -> Dict[str, List[str]]:
    index: Dict[str, List[str]] = defaultdict(list)
    for org_id, name in conn.execute('SELECT id, "이름" FROM organization'):
        name_clean = clean_str(name)
        if not name_clean:
            continue
        index[name_clean].append(org_id)
    return index


def load_people_by_org(conn: sqlite3.Connection) -> Dict[str, List[sqlite3.Row]]:
    """
    Load all people once (no index on organizationId), grouped by organizationId.
    """
    people_by_org: Dict[str, List[sqlite3.Row]] = defaultdict(list)
    cur = conn.execute('SELECT organizationId,"RecordId","이름","담당자","소속 상위 조직" FROM people')
    for row in cur:
        people_by_org[row["organizationId"]].append(row)
    return people_by_org


def build_name_to_team(part_structure: Dict[str, Dict[str, Iterable[str]]]) -> Dict[str, List[str]]:
    mapping: Dict[str, List[str]] = {}
    for team, parts in part_structure.items():
        for names in parts.values():
            for name in names:
                key = name.strip()
                if not key:
                    continue
                mapping.setdefault(key, []).append(team)
    return mapping


def process_target_sheet(sheet_name: str, output_path: Path) -> dict:
    if SOURCE_WB is None or ORG_INDEX is None or PEOPLE_BY_ORG is None or NAME_TO_TEAM is None:
        raise RuntimeError("Context not initialized. Run main().")

    wb = SOURCE_WB
    result_rows: List[Tuple[str, str, str, str, str]] = []
    log_rows: List[Tuple[object, ...]] = []

    skip_counts = defaultdict(int)
    team_not_found = 0
    team_ambiguous = 0
    owner_json_candidate_count = 0
    owner_json_extracted_count = 0
    owner_json_parse_failed_count = 0
    owner_json_parse_errors: List[str] = []

    def append_log(
        excel_row_index: int,
        org_name: str,
        upper_org_raw: object,
        assignee_raw: object,
        assignee_tokens_count: int,
        org_lookup_status: str,
        matched_people_count: int,
        reason: str,
        owner_raw: object = "",
        owner_name_normalized: str = "",
        owner_team: str = "",
    ) -> None:
        log_rows.append(
            (
                sheet_name,
                excel_row_index,
                org_name,
                clean_str(upper_org_raw),
                clean_str(assignee_raw),
                assignee_tokens_count,
                org_lookup_status,
                matched_people_count,
                reason,
                clean_str(owner_raw),
                owner_name_normalized,
                owner_team,
            )
        )

    if sheet_name not in wb.sheetnames:
        skip_counts["sheet_missing"] += 1
        append_log(
            excel_row_index=0,
            org_name="",
            upper_org_raw="",
            assignee_raw="",
            assignee_tokens_count=0,
            org_lookup_status="not_applicable",
            matched_people_count=0,
            reason="sheet_missing",
        )
    else:
        ws = wb[sheet_name]
        for excel_row in range(2, ws.max_row + 1):
            org_name = clean_str(ws.cell(excel_row, 2).value)
            upper_org_raw = ws.cell(excel_row, 3).value
            assignee_cell = ws.cell(excel_row, 4).value

            assignee_value, assignee_err, assignee_tokens_count = parse_assignee(assignee_cell)
            upper_org_norm = norm_upper_org(upper_org_raw)

            if assignee_err == "assignee_multi":
                skip_counts[assignee_err] += 1
                append_log(
                    excel_row,
                    org_name,
                    upper_org_raw,
                    assignee_cell,
                    assignee_tokens_count,
                    "not_found",
                    0,
                    assignee_err,
                )
                continue
            if assignee_err == "assignee_empty":
                skip_counts[assignee_err] += 1
                append_log(
                    excel_row,
                    org_name,
                    upper_org_raw,
                    assignee_cell,
                    assignee_tokens_count,
                    "not_found",
                    0,
                    assignee_err,
                )
                continue

            if not org_name:
                skip_counts["org_not_found"] += 1
                append_log(
                    excel_row,
                    org_name,
                    upper_org_raw,
                    assignee_cell,
                    assignee_tokens_count,
                    "not_found",
                    0,
                    "org_not_found",
                )
                continue

            org_ids = ORG_INDEX.get(org_name, [])
            if not org_ids:
                skip_counts["org_not_found"] += 1
                append_log(
                    excel_row,
                    org_name,
                    upper_org_raw,
                    assignee_cell,
                    assignee_tokens_count,
                    "not_found",
                    0,
                    "org_not_found",
                )
                continue
            if len(org_ids) > 1:
                skip_counts["org_ambiguous"] += 1
                append_log(
                    excel_row,
                    org_name,
                    upper_org_raw,
                    assignee_cell,
                    assignee_tokens_count,
                    "ambiguous",
                    0,
                    "org_ambiguous",
                )
                continue

            org_id = org_ids[0]
            people = PEOPLE_BY_ORG.get(org_id, [])

            matched_people = [
                person
                for person in people
                if norm_upper_org(person["소속 상위 조직"]) == upper_org_norm
            ]

            if not matched_people:
                skip_counts["people_zero_match"] += 1
                append_log(
                    excel_row,
                    org_name,
                    upper_org_raw,
                    assignee_cell,
                    assignee_tokens_count,
                    "found",
                    0,
                    "people_zero_match",
                )
                continue

            for person in matched_people:
                existing_owner_raw = person["담당자"]
                owner_name, owner_json_candidate, owner_json_extracted, owner_json_error = normalize_owner_name(
                    existing_owner_raw
                )

                if owner_json_candidate:
                    owner_json_candidate_count += 1
                if owner_json_extracted:
                    owner_json_extracted_count += 1
                if owner_json_error:
                    owner_json_parse_failed_count += 1
                    if len(owner_json_parse_errors) < 5:
                        owner_json_parse_errors.append(owner_json_error)

                teams = NAME_TO_TEAM.get(owner_name, [])
                team_name = ""
                if not teams:
                    team_not_found += 1
                    append_log(
                        excel_row,
                        org_name,
                        upper_org_raw,
                        assignee_cell,
                        assignee_tokens_count,
                        "found",
                        len(matched_people),
                        "team_not_found",
                        owner_raw=existing_owner_raw,
                        owner_name_normalized=owner_name,
                        owner_team="",
                    )
                elif len(teams) == 1:
                    team_name = teams[0]
                else:
                    team_ambiguous += 1
                    team_name = ",".join(sorted(set(teams)))
                    append_log(
                        excel_row,
                        org_name,
                        upper_org_raw,
                        assignee_cell,
                        assignee_tokens_count,
                        "found",
                        len(matched_people),
                        "team_ambiguous",
                        owner_raw=existing_owner_raw,
                        owner_name_normalized=owner_name,
                        owner_team=team_name,
                    )

                result_rows.append(
                    (
                        clean_str(person["RecordId"]),
                        clean_str(person["이름"]),
                        assignee_value or "",
                        owner_name,
                        team_name,
                    )
                )

    out_wb = Workbook()
    res_ws = out_wb.active
    res_ws.title = RESULT_SHEET
    res_ws.append(
        [
            "People - RecordID",
            "People - 이름",
            "People - 담당자",
            "People - 담당자(기존)",
            "People - 담당자(기존) 소속 팀",
        ]
    )
    for row in result_rows:
        res_ws.append(row)

    log_ws = out_wb.create_sheet(LOG_SHEET)
    log_ws.append(
        [
            "sheet_name",
            "excel_row_index",
            "org_name",
            "upper_org",
            "assignee_raw",
            "assignee_tokens_count",
            "org_lookup_status",
            "matched_people_count",
            "reason",
            "owner_raw",
            "owner_name_normalized",
            "owner_team",
        ]
    )
    for row in log_rows:
        log_ws.append(list(row))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)

    skipped_total = sum(skip_counts.values())
    team_unmatched_count = team_not_found + team_ambiguous

    return {
        "sheet_name": sheet_name,
        "result_rows": len(result_rows),
        "skipped_total": skipped_total,
        "skipped_by_reason": dict(skip_counts),
        "team_unmatched_count": team_unmatched_count,
        "team_not_found": team_not_found,
        "team_ambiguous": team_ambiguous,
        "org_not_found": skip_counts.get("org_not_found", 0),
        "org_ambiguous": skip_counts.get("org_ambiguous", 0),
        "assignee_multi": skip_counts.get("assignee_multi", 0),
        "assignee_empty": skip_counts.get("assignee_empty", 0),
        "people_zero_match": skip_counts.get("people_zero_match", 0),
        "output_path": output_path,
        "owner_json_candidate": owner_json_candidate_count,
        "owner_json_extracted": owner_json_extracted_count,
        "owner_json_parse_failed": owner_json_parse_failed_count,
        "owner_json_parse_errors": owner_json_parse_errors,
    }


def main() -> None:
    global SOURCE_WB, ORG_INDEX, PEOPLE_BY_ORG, NAME_TO_TEAM

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    ORG_INDEX = build_org_index(conn)
    PEOPLE_BY_ORG = load_people_by_org(conn)
    NAME_TO_TEAM = build_name_to_team(PART_STRUCTURE)
    SOURCE_WB = load_workbook(EXCEL_PATH)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    summaries = []
    for sheet_name, filename in SHEETS:
        output_path = OUTPUT_DIR / filename
        summary = process_target_sheet(sheet_name, output_path)
        summaries.append(summary)

    for summary in summaries:
        print(f"[{summary['sheet_name']}]")
        print(f"  result rows: {summary['result_rows']}")
        print(
            "  skipped total: {total} (by reason: {by_reason})".format(
                total=summary["skipped_total"], by_reason=summary["skipped_by_reason"]
            )
        )
        print(
            "  team unmatched: {count} (not_found={nf}, ambiguous={amb})".format(
                count=summary["team_unmatched_count"], nf=summary["team_not_found"], amb=summary["team_ambiguous"]
            )
        )
        print(f"  output file: {summary['output_path']}")
        print(
            "  owner JSON stats: candidate={cand}, extracted={ext}, parse_failed={pf}".format(
                cand=summary["owner_json_candidate"],
                ext=summary["owner_json_extracted"],
                pf=summary["owner_json_parse_failed"],
            )
        )
        if summary["owner_json_parse_errors"]:
            print("    sample errors: " + "; ".join(summary["owner_json_parse_errors"]))


if __name__ == "__main__":
    main()
