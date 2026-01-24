import tempfile
import unittest
from pathlib import Path

import openpyxl

from dashboard.server import counterparty_targets_2026 as ct


class CounterpartyTargetsLoaderTest(unittest.TestCase):
    def _write_wb(self, path: Path, rows: list[tuple[str, str, float]]) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "26 출강 타겟"
        ws.append(["기업명", "카운터파티", "26 출강 타겟"])
        for r in rows:
            ws.append(list(r))
        ws2 = wb.create_sheet("26 온라인 타겟")
        ws2.append(["기업명", "카운터파티", "26 온라인 타겟"])
        ws2.append(["ORG-A", "CP-A", 1.0])
        wb.save(path)

    def test_loader_parses_and_converts_and_skips_duplicates(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "counterparty_targets_2026.xlsx"
            # duplicate key for offline should be skipped
            self._write_wb(
                xlsx_path,
                [
                    ("ORG-A", "CP-A", 2.16),
                    ("ORG-A", "CP-A", 3.0),
                    ("ORG-B", "-", 1.0),
                ],
            )
            # reset cache and path
            original_path = ct.RESOURCE_PATH
            try:
                ct._CACHE.update({"mtime": None, "offline": {}, "online": {}, "meta": {}})
                ct.RESOURCE_PATH = xlsx_path

                offline, online, meta, version = ct.load_counterparty_targets_2026()
                # duplicate key skipped
                self.assertNotIn(("ORG-A", "CP-A"), offline)
                # dash upper -> 미입력, multiplied to won
                self.assertEqual(offline[("ORG-B", "미입력")], 1.0 * 1e8)
                self.assertEqual(online[("ORG-A", "CP-A")], 1.0 * 1e8)
                self.assertIn(("ORG-B", "미입력"), meta)
                self.assertEqual(meta[("ORG-B", "미입력")]["orgName"], "ORG-B")
                self.assertTrue(version.startswith("xlsx_mtime:"))
            finally:
                ct.RESOURCE_PATH = original_path
                ct._CACHE.update({"mtime": None, "offline": {}, "online": {}, "meta": {}})

    def test_loader_supports_org_id_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "counterparty_targets_2026.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "26 출강 타겟"
            ws.append(["기업명", "orgId", "카운터파티", "26 출강 타겟"])
            ws.append(["ORG-ID-NAME", "org-123", "Upper Team", 2.0])
            ws2 = wb.create_sheet("26 온라인 타겟")
            ws2.append(["기업명", "orgId", "카운터파티", "26 온라인 타겟"])
            ws2.append(["ORG-ID-NAME", "org-123", "Upper Team", 3.0])
            wb.save(xlsx_path)

            original_path = ct.RESOURCE_PATH
            try:
                ct._CACHE.update({"mtime": None, "offline": {}, "online": {}, "meta": {}})
                ct.RESOURCE_PATH = xlsx_path
                offline, online, meta, _ = ct.load_counterparty_targets_2026()
            finally:
                ct.RESOURCE_PATH = original_path
                ct._CACHE.update({"mtime": None, "offline": {}, "online": {}, "meta": {}})

            self.assertEqual(offline[("ORG-ID-NAME", "Upper Team")], 2.0 * 1e8)
            self.assertEqual(online[("ORG-ID-NAME", "Upper Team")], 3.0 * 1e8)
            self.assertEqual(meta[("ORG-ID-NAME", "Upper Team")]["orgId"], "org-123")


if __name__ == "__main__":
    unittest.main()
